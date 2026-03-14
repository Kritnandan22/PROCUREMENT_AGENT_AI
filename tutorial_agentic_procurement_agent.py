"""
Tutorial-defined Agentic Procurement AI
======================================

Implements the workflows described in TUTORIAL_Agentic_Procurement_AI.md as a
standalone, terminal-friendly agent that can be run without modifying the
existing procurement agent.

Design principles copied from the tutorial:
  - Read Oracle data safely through a constrained query layer
  - Start with Level 0-1 autonomy by default
  - Allow Level 2 behaviour only as draft PO payload generation
  - Keep every decision auditable with rationale and evidence
  - Separate data access, decision logic, and action generation

This script does NOT write back to Oracle. Draft POs are emitted as structured
payloads for buyer review, matching the tutorial's cautious write-back model.

Run examples:
  python tutorial_agentic_procurement_agent.py --workflow exception-triage
    python tutorial_agentic_procurement_agent.py \
            --workflow all --autonomy-level 2
"""

from __future__ import annotations

import argparse
import json
import os
import traceback
from collections import Counter
from contextlib import contextmanager
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import anthropic
import oracledb
import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill

from config import get_config


ROOT_DIR = Path(__file__).resolve().parent
TUTORIAL_PATH = ROOT_DIR / "TUTORIAL_Agentic_Procurement_AI.md"

# Load configuration (from config.yaml + environment variables)
_app_config = get_config()
OUTPUT_DIR = _app_config.output_dir

# Legacy compatibility - expose config values as module-level variables
DB_HOST = _app_config.database.host
DB_PORT = _app_config.database.port
DB_SID = _app_config.database.sid
DB_SERVICE_NAME = _app_config.database.service_name
APPS_USER = _app_config.database.user
APPS_PASSWORD = _app_config.database.password
ORACLE_CLIENT_PATH = _app_config.database.oracle_client_path


def _init_oracle_client() -> bool:
    try:
        if ORACLE_CLIENT_PATH:
            oracledb.init_oracle_client(lib_dir=ORACLE_CLIENT_PATH)
        else:
            oracledb.init_oracle_client()
        return True
    except Exception as e:
        import traceback
        print(f"FAILED TO INIT ORACLE THICK MODE: {e}")
        traceback.print_exc()
        return False


THICK_MODE = _init_oracle_client()


def _build_dsn() -> str:
    if DB_SID:
        return oracledb.makedsn(DB_HOST, DB_PORT, sid=DB_SID)
    if DB_SERVICE_NAME:
        return oracledb.makedsn(DB_HOST, DB_PORT, service_name=DB_SERVICE_NAME)
    raise ValueError("Set DB_SID or DB_SERVICE_NAME in .env")


DSN = _build_dsn()
_pool: oracledb.ConnectionPool | None = None


def _get_pool() -> oracledb.ConnectionPool:
    global _pool
    if _pool is None:
        _pool = oracledb.create_pool(
            user=APPS_USER,
            password=APPS_PASSWORD,
            dsn=DSN,
            min=1,
            max=4,
            increment=1,
            stmtcachesize=50,
            getmode=oracledb.POOL_GETMODE_WAIT,
        )
    return _pool


@contextmanager
def _conn() -> Any:
    pool = _get_pool()
    conn = pool.acquire()
    try:
        yield conn
    finally:
        pool.release(conn)


def _rows_to_dicts(cursor: oracledb.Cursor) -> list[dict[str, Any]]:
    if cursor.description is None:
        return []
    columns = [desc[0].lower() for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _json_default(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _read_only_sql(sql: str) -> str:
    sql_clean = sql.strip().upper()
    if not sql_clean.startswith("SELECT") and not sql_clean.startswith("WITH"):
        raise ValueError("Only SELECT and WITH queries are allowed.")
    return sql


@dataclass
class ActionRecord:
    workflow: str
    action: str
    priority: str
    requires_human_approval: bool
    summary: str
    rationale: str
    evidence: dict[str, Any]
    payload: dict[str, Any]
    confidence: float | None = None
    assigned_to: str | None = None


class ProcurementDecisionEngine:
    """Internal rules-based decision engine for complex procurement decisions.

    No external API dependency. Uses weighted multi-factor scoring to evaluate
    scenarios and produce confidence-scored recommendations with rationale.

    All thresholds and parameters are loaded from configuration for production flexibility.

    Decision Categories:
    - supplier_switch: Should we switch to an alternate supplier?
    - price_renegotiation: Should we renegotiate pricing?
    - consolidation: Should we consolidate spend to one supplier?
    - escalation: What priority level does this situation warrant?
    - po_creation: Should a draft PO be created automatically?
    """

    def __init__(self):
        """Initialize decision engine with configuration-based thresholds."""
        self.config = _app_config
        self.de_config = self.config.decision_engine
        self.priorities = self.config.get_priorities()

    @property
    def _PRIORITY_SCORES(self) -> dict[str, float]:
        """Get priority scores from configuration."""
        return self.priorities

    def decide_supplier_switch(
        self,
        current_vendor: dict[str, Any],
        current_performance: dict[str, Any],
        alternates: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Decide whether to recommend switching to an alternate supplier.

        Factors: on-time rate, lead time, number of late deliveries, alternate availability.
        Returns: decision (switch/monitor/escalate), confidence, rationale, recommended_vendor.
        """
        on_time_rate = float(
            current_performance.get("on_time_rate_pct") or 100)
        total_deliveries = int(
            current_performance.get("total_deliveries") or 0)
        total_late_days = int(current_performance.get("total_late_days") or 0)

        # Score: lower on-time rate = higher urgency to switch
        # 0.0 (perfect) to 1.0 (never on-time)
        performance_score = (100 - on_time_rate) / 100
        # more deliveries = more statistical confidence
        volume_score = min(total_deliveries / self.de_config.min_deliveries, 1.0)
        # avg days late per delivery weighted
        lateness_severity = min(total_late_days / self.de_config.on_time_rate_switch, 1.0)

        switch_score = (performance_score * 0.5) + \
            (volume_score * 0.3) + (lateness_severity * 0.2)
        has_viable_alternate = bool(alternates)

        if switch_score >= 0.6 and on_time_rate < self.de_config.on_time_rate_switch and has_viable_alternate:
            decision = "switch"
            confidence = round(min(switch_score, 1.0), 2)
            best_alt = alternates[0]
            rationale = (
                f"Supplier {current_vendor.get('vendor_name', 'UNKNOWN')} has "
                f"{on_time_rate:.0f}% on-time delivery over {total_deliveries} orders. "
                f"Switch recommended to {best_alt.get('vendor_name', 'alternate')} "
                f"(lead time: {best_alt.get('lead_time', 'N/A')} days)."
            )
            recommended_vendor = best_alt
        elif switch_score >= 0.4 or on_time_rate < self.de_config.on_time_rate_monitor:
            decision = "monitor"
            confidence = round(switch_score, 2)
            rationale = (
                f"Supplier {current_vendor.get('vendor_name', 'UNKNOWN')} shows "
                f"{on_time_rate:.0f}% on-time rate (target: {self.de_config.on_time_rate_monitor}%). "
                f"Below target but insufficient history for switch recommendation. Escalate if trend continues."
            )
            recommended_vendor = None
        else:
            decision = "no_action"
            confidence = round(1 - switch_score, 2)
            rationale = (
                f"Supplier {current_vendor.get('vendor_name', 'UNKNOWN')} performing "
                f"acceptably at {on_time_rate:.0f}% on-time rate."
            )
            recommended_vendor = None

        return {
            "decision": decision,
            "confidence": confidence,
            "rationale": rationale,
            "switch_score": round(switch_score, 3),
            "on_time_rate_pct": on_time_rate,
            "total_deliveries": total_deliveries,
            "recommended_vendor": recommended_vendor,
            "has_viable_alternate": has_viable_alternate,
        }

    def decide_price_renegotiation(
        self,
        item_id: int,
        current_price: float,
        historical_avg: float,
        contract_price: float | None,
        annual_volume: float,
    ) -> dict[str, Any]:
        """Decide whether and how urgently to renegotiate price.

        Factors: deviation from historical avg, deviation from contract, potential savings.
        Returns: decision, confidence, savings estimate, renegotiation strategy.
        """
        if historical_avg <= 0:
            return {"decision": "insufficient_data", "confidence": 0.0, "rationale": "No price history."}

        hist_deviation = (current_price - historical_avg) / \
            historical_avg * 100
        contract_deviation = None
        if contract_price and contract_price > 0:
            contract_deviation = (
                current_price - contract_price) / contract_price * 100

        # Score based on how far above baseline
        price_urgency = min(abs(hist_deviation) / 50, 1.0)  # 50% = max urgency
        savings_potential = (current_price - historical_avg) * \
            annual_volume if hist_deviation > 0 else 0

        if contract_deviation is not None and contract_deviation > self.de_config.contract_breach_threshold:
            decision = "renegotiate_immediately"
            confidence = min(0.7 + (contract_deviation / 100), 1.0)
            strategy = "enforce_contract"
            rationale = (
                f"Item {item_id} priced at ${current_price:.2f}, which is "
                f"{contract_deviation:.1f}% above active contract price ${contract_price:.2f}. "
                f"Exceeds threshold ({self.de_config.contract_breach_threshold}%). Immediate enforcement required. "
                f"Potential savings: ${savings_potential:,.2f}/year."
            )
        elif hist_deviation > self.de_config.price_dev_renegotiate:
            decision = "renegotiate"
            confidence = round(price_urgency, 2)
            strategy = "historical_baseline"
            rationale = (
                f"Item {item_id} priced at ${current_price:.2f}, which is "
                f"{hist_deviation:.1f}% above 12-month average ${historical_avg:.2f}. "
                f"Exceeds threshold ({self.de_config.price_dev_renegotiate}%). Recommend renegotiation. "
                f"Estimated savings: ${savings_potential:,.2f}/year."
            )
        elif hist_deviation > self.de_config.price_dev_review:
            decision = "review"
            confidence = round(price_urgency * 0.7, 2)
            strategy = "monitor_and_negotiate"
            rationale = (
                f"Item {item_id} showing {hist_deviation:.1f}% price increase vs baseline. "
                f"Monitor and include in next negotiation cycle."
            )
        else:
            decision = "no_action"
            confidence = round(1 - price_urgency, 2)
            strategy = "none"
            rationale = f"Item {item_id} price within acceptable range vs historical baseline."

        return {
            "decision": decision,
            "confidence": confidence,
            "rationale": rationale,
            "hist_deviation_pct": round(hist_deviation, 2),
            "contract_deviation_pct": round(contract_deviation, 2) if contract_deviation is not None else None,
            "estimated_savings": round(savings_potential, 2),
            "renegotiation_strategy": strategy,
        }

    def decide_consolidation(
        self,
        item_id: int,
        supplier_count: int,
        min_price: float,
        max_price: float,
        total_spend: float,
        po_count: int,
    ) -> dict[str, Any]:
        """Decide if spend consolidation is worthwhile for a multi-sourced item.

        Factors: price spread, number of suppliers, total spend volume.
        Returns: decision, confidence, estimated savings, consolidation plan.
        """
        if min_price <= 0 or supplier_count < 2:
            return {"decision": "no_action", "confidence": 0.0, "rationale": "Insufficient data."}

        spread_pct = (max_price - min_price) / min_price * 100
        savings_if_consolidated = (
            max_price - min_price) * (total_spend / max_price) if max_price > 0 else 0

        # Score: high spread + high volume = consolidate
        spread_urgency = min(spread_pct / 30, 1.0)  # 30% spread = max urgency
        volume_factor = min(total_spend / 100000, 1.0)  # $100k = high volume
        consolidation_score = (spread_urgency * 0.6) + (volume_factor * 0.4)

        if consolidation_score >= 0.6 and spread_pct > 15:
            decision = "consolidate"
            confidence = round(consolidation_score, 2)
            rationale = (
                f"Item {item_id} purchased from {supplier_count} suppliers with "
                f"{spread_pct:.1f}% price spread. Consolidating to lowest-price supplier "
                f"could save ~${savings_if_consolidated:,.2f}."
            )
        elif spread_pct > 10 and total_spend > 10000:
            decision = "review"
            confidence = round(consolidation_score * 0.7, 2)
            rationale = (
                f"Item {item_id} has {spread_pct:.1f}% price spread across "
                f"{supplier_count} suppliers. Review in next sourcing cycle."
            )
        else:
            decision = "no_action"
            confidence = round(1 - consolidation_score, 2)
            rationale = f"Item {item_id} multi-sourcing with acceptable spread ({spread_pct:.1f}%)."

        return {
            "decision": decision,
            "confidence": confidence,
            "rationale": rationale,
            "spread_pct": round(spread_pct, 2),
            "consolidation_score": round(consolidation_score, 3),
            "estimated_savings": round(savings_if_consolidated, 2),
            "recommended_action": f"Route all POs for item {item_id} to lowest-price supplier"
            if decision == "consolidate" else "Monitor",
        }

    def score_exception_priority(
        self,
        exception_type: int,
        quantity: float,
        days_overdue: int = 0,
        unit_price: float = 0.0,
        pegging_count: int = 0,
    ) -> dict[str, Any]:
        """Score an exception to determine priority and recommended action.

        Multi-factor scoring: exception type severity + financial impact + days overdue.
        Returns: priority, score, action, confidence.
        """
        # Type severity weights (higher = more critical)
        TYPE_SEVERITY = {
            6: 0.9,   # Past Due Purchase Orders - very critical
            3: 0.8,   # Items Below Safety Stock - high
            13: 0.85,  # Demand Not Satisfied - critical
            12: 0.7,  # Demand Past Due - high
            2: 0.75,  # Late Replenishment - high
            11: 0.6,  # Orders at Risk - medium-high
            7: 0.5,   # Cancelled Supply
            8: 0.6,   # Short Shipments
            4: 0.4,   # Excess Inventory
            1: 0.3,   # Items with No Activity - low
        }
        type_severity = TYPE_SEVERITY.get(exception_type, 0.5)

        # Financial impact score
        financial_impact = (
            quantity * unit_price) if unit_price > 0 else quantity
        # $50k = max financial urgency
        financial_score = min(financial_impact / 50000, 1.0)

        # Time urgency score
        # 30 days = max overdue
        time_score = min(days_overdue / 30, 1.0) if days_overdue > 0 else 0.3
        # more pegging = more demand at risk
        demand_coverage_score = min(pegging_count / 10, 1.0)

        # Composite score
        composite = (
            type_severity * 0.4
            + financial_score * 0.25
            + time_score * 0.2
            + demand_coverage_score * 0.15
        )

        if composite >= 0.75:
            priority = "P1-CRITICAL"
            action = "recommend_draft_po"
        elif composite >= 0.55:
            priority = "P2-HIGH"
            action = "recommend_draft_po"
        elif composite >= 0.35:
            priority = "P3-MEDIUM"
            action = "notify_buyer"
        else:
            priority = "P4-LOW"
            action = "report_insight"

        return {
            "priority": priority,
            "action": action,
            "composite_score": round(composite, 3),
            "confidence": round(min(composite + 0.1, 1.0), 2),
            "factors": {
                "type_severity": type_severity,
                "financial_score": round(financial_score, 3),
                "time_urgency": round(time_score, 3),
                "demand_coverage": round(demand_coverage_score, 3),
            },
        }

    def analyze_po_risk(
        self,
        quantity: float,
        unit_price: float,
        lead_time_days: int,
        need_by_date_days: int,
        supplier_on_time_rate: float,
    ) -> dict[str, Any]:
        """Assess risk of a proposed PO and recommend safeguards.

        Returns: risk_level, confidence, risk_factors, safeguards.
        """
        financial_risk = min((quantity * unit_price) / 100000, 1.0)
        time_risk = 1.0 if lead_time_days > need_by_date_days else max(
            (lead_time_days / max(need_by_date_days, 1)), 0
        )
        supplier_risk = (100 - supplier_on_time_rate) / 100

        risk_score = (financial_risk * 0.35) + \
            (time_risk * 0.4) + (supplier_risk * 0.25)

        if risk_score >= 0.7:
            risk_level = "HIGH"
        elif risk_score >= 0.4:
            risk_level = "MEDIUM"
        else:
            risk_level = "LOW"

        safeguards = []
        if time_risk >= 0.9:
            safeguards.append(
                "URGENT: Lead time exceeds need-by date - consider alternate supplier")
        if supplier_on_time_rate < 70:
            safeguards.append(
                "Supplier reliability is low - add buffer stock or dual-source")
        if financial_risk > 0.5:
            safeguards.append(
                "High-value PO - require manager approval before issuing")
        if not safeguards:
            safeguards.append("Standard procurement process applicable")

        return {
            "risk_level": risk_level,
            "risk_score": round(risk_score, 3),
            "confidence": round(0.85, 2),
            "risk_factors": {
                "financial_risk": round(financial_risk, 3),
                "time_risk": round(time_risk, 3),
                "supplier_reliability_risk": round(supplier_risk, 3),
            },
            "safeguards": safeguards,
        }


class OracleReadOnlyGateway:
    """Read-only Oracle access layer.

    This mirrors the tutorial's MCP safety boundary inside a local runner.
    All table names and business logic parameters are loaded from configuration.
    """

    def __init__(self):
        """Initialize gateway with configuration."""
        self.config = _app_config
        self.tables = self.config.tables
        self.exception_types = self.config.get_exception_types()

    def test_connection(self) -> dict[str, Any]:
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT VERSION, USER, SYSDATE FROM V$INSTANCE, DUAL"
                )
                version, user_name, sysdate = cur.fetchone()
        return {
            "db_version": version,
            "connected_user": user_name,
            "db_sysdate": sysdate,
            "host": DB_HOST,
            "port": DB_PORT,
            "sid": DB_SID or DB_SERVICE_NAME,
            "thick_mode": THICK_MODE,
        }

    def execute_query(
        self,
        sql: str,
        binds: dict[str, Any] | None = None,
        max_rows: int = 200,
    ) -> list[dict[str, Any]]:
        query = _read_only_sql(sql)
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(query, binds or {})
                rows = _rows_to_dicts(cur)
        return rows[:max_rows]

    def list_tables(
        self,
        schema: str,
        name_filter: str = "",
    ) -> list[dict[str, Any]]:
        sql = """
            SELECT OWNER, TABLE_NAME, NUM_ROWS, LAST_ANALYZED
            FROM ALL_TABLES
            WHERE OWNER = :owner
        """
        binds: dict[str, Any] = {"owner": schema.upper()}
        if name_filter:
            sql += " AND TABLE_NAME LIKE :name_filter"
            binds["name_filter"] = f"%{name_filter.upper()}%"
        sql += " ORDER BY TABLE_NAME"
        return self.execute_query(sql, binds=binds, max_rows=500)

    def get_exception_summary(self, limit: int = 10) -> dict[str, Any]:
        total_rows = self.execute_query(
            (
                f"SELECT COUNT(*) AS total_exceptions "
                f"FROM {self.tables.msc_exception_details}"
            ),
            max_rows=1,
        )
        top_plans = self.execute_query(
            f"""
            SELECT *
            FROM (
                SELECT p.PLAN_ID,
                       p.COMPILE_DESIGNATOR AS plan_name,
                       COUNT(e.EXCEPTION_DETAIL_ID) AS exception_count
                FROM {self.tables.msc_plans} p
                JOIN {self.tables.msc_exception_details} e ON e.PLAN_ID = p.PLAN_ID
                GROUP BY p.PLAN_ID, p.COMPILE_DESIGNATOR
                ORDER BY COUNT(e.EXCEPTION_DETAIL_ID) DESC
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"limit": limit},
            max_rows=limit,
        )
        return {
            "total_exceptions": (
                total_rows[0]["total_exceptions"] if total_rows else 0
            ),
            "top_plans": top_plans,
        }

    def get_exception_types(self, plan_id: int) -> list[dict[str, Any]]:
        return self.execute_query(
            f"""
            SELECT e.EXCEPTION_TYPE,
                   COUNT(*) AS exception_count,
                   ROUND(AVG(NVL(e.QUANTITY, 0)), 2) AS avg_quantity,
                   ROUND(SUM(NVL(e.QUANTITY, 0)), 2) AS total_quantity
            FROM {self.tables.msc_exception_details} e
            WHERE e.PLAN_ID = :plan_id
            GROUP BY e.EXCEPTION_TYPE
            ORDER BY COUNT(*) DESC
            """,
            binds={"plan_id": plan_id},
            max_rows=100,
        )

    def get_exception_details(
        self,
        plan_id: int,
        exception_type: int,
        limit: int,
    ) -> list[dict[str, Any]]:
        return self.execute_query(
            f"""
            SELECT *
            FROM (
                SELECT e.EXCEPTION_DETAIL_ID,
                       e.PLAN_ID,
                       e.EXCEPTION_TYPE,
                       e.INVENTORY_ITEM_ID,
                       e.ORGANIZATION_ID,
                       NVL(e.QUANTITY, 0) AS quantity,
                       e.DATE1,
                       e.DATE2
                FROM {self.tables.msc_exception_details} e
                WHERE e.PLAN_ID = :plan_id
                  AND e.EXCEPTION_TYPE = :exception_type
                ORDER BY NVL(e.QUANTITY, 0) DESC, e.EXCEPTION_DETAIL_ID
            )
            WHERE ROWNUM <= :limit
            """,
            binds={
                "plan_id": plan_id,
                "exception_type": exception_type,
                "limit": limit,
            },
            max_rows=limit,
        )

    def get_item_context(self, item_id: int) -> dict[str, Any]:
        rows = self.execute_query(
            f"""
            SELECT *
            FROM (
                SELECT INVENTORY_ITEM_ID,
                       SEGMENT1 AS item_number,
                       DESCRIPTION,
                       PRIMARY_UOM_CODE,
                       LIST_PRICE_PER_UNIT,
                       FULL_LEAD_TIME,
                       BUYER_ID,
                       PURCHASING_ENABLED_FLAG
                FROM {self.tables.system_items_b}
                WHERE INVENTORY_ITEM_ID = :item_id
            )
            WHERE ROWNUM <= 1
            """,
            binds={"item_id": item_id},
            max_rows=1,
        )
        return rows[0] if rows else {}

    def get_supplier_options(
        self,
        item_id: int,
        limit: int = 5,
    ) -> list[dict[str, Any]]:
        """Get approved suppliers for an item, sorted cheapest+fastest (tutorial spec: Workflow 3)."""
        try:
            return self.execute_query(
                """
                SELECT *
                FROM (
                    SELECT mis.INVENTORY_ITEM_ID,
                           mis.VENDOR_ID,
                           pv.VENDOR_NAME,
                           mis.VENDOR_SITE_ID,
                           pvs.VENDOR_SITE_CODE,
                           NVL(mis.LEAD_TIME, 999)            AS lead_time_days,
                           NVL(mis.MIN_ORDER_QTY, 0)          AS min_order_qty,
                           NVL(pvs.LEAD_TIME_DAYS, mis.LEAD_TIME) AS site_lead_time,
                           NVL(pvs.PRICE_OVERRIDE, 0)         AS list_price
                    FROM MSC.MSC_ITEM_SUPPLIERS mis
                    LEFT JOIN APPS.PO_VENDORS pv
                           ON pv.VENDOR_ID = mis.VENDOR_ID
                    LEFT JOIN APPS.PO_VENDOR_SITES_ALL pvs
                           ON pvs.VENDOR_SITE_ID = mis.VENDOR_SITE_ID
                           AND pvs.INACTIVE_DATE IS NULL
                    WHERE mis.INVENTORY_ITEM_ID = :item_id
                    AND (mis.DISABLE_DATE IS NULL OR mis.DISABLE_DATE > TRUNC(SYSDATE))
                    ORDER BY NVL(mis.LEAD_TIME, 999) ASC,
                             NVL(pvs.PRICE_OVERRIDE, 0) ASC,
                             pv.VENDOR_NAME ASC
                )
                WHERE ROWNUM <= :limit
                """,
                binds={"item_id": item_id, "limit": limit},
                max_rows=limit,
            )
        except Exception:
            return self.execute_query(
                """
                SELECT *
                FROM (
                    SELECT VENDOR_ID, VENDOR_NAME
                    FROM APPS.PO_VENDORS
                    ORDER BY VENDOR_NAME
                )
                WHERE ROWNUM <= :limit
                """,
                binds={"limit": limit},
                max_rows=limit,
            )

    def get_open_po_coverage(
        self,
        item_id: int,
        limit: int = 10,
    ) -> list[dict[str, Any]]:
        return self.execute_query(
            """
            SELECT *
            FROM (
                SELECT ph.SEGMENT1 AS po_number,
                       ph.VENDOR_ID,
                       ph.AUTHORIZATION_STATUS,
                       pll.NEED_BY_DATE,
                       (
                           pll.QUANTITY
                           - NVL(pll.QUANTITY_RECEIVED, 0)
                       ) AS qty_outstanding
                FROM {self.tables.po_headers_all} ph
                JOIN {self.tables.po_lines_all} pl ON pl.PO_HEADER_ID = ph.PO_HEADER_ID
                JOIN APPS.PO_LINE_LOCATIONS_ALL pll
                  ON pll.PO_LINE_ID = pl.PO_LINE_ID
                WHERE pl.ITEM_ID = :item_id
                  AND ph.TYPE_LOOKUP_CODE = 'STANDARD'
                  AND ph.AUTHORIZATION_STATUS IN (
                      'APPROVED',
                      'IN PROCESS',
                      'INCOMPLETE'
                  )
                  AND NVL(pll.QUANTITY_RECEIVED, 0)
                      < NVL(pll.QUANTITY, 0)
                ORDER BY pll.NEED_BY_DATE NULLS LAST
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"item_id": item_id, "limit": limit},
            max_rows=limit,
        )

    def get_safety_stock_context(
        self,
        item_id: int,
        plan_id: int,
    ) -> dict[str, Any]:
        rows = self.execute_query(
            """
            SELECT *
            FROM (
                SELECT ss.PLAN_ID,
                       ss.INVENTORY_ITEM_ID,
                       ss.SAFETY_STOCK_QUANTITY,
                       NVL(s.current_supply, 0) AS current_supply,
                       (
                           ss.SAFETY_STOCK_QUANTITY
                           - NVL(s.current_supply, 0)
                       ) AS shortage
                FROM MSC.MSC_SAFETY_STOCKS ss
                LEFT JOIN (
                    SELECT PLAN_ID,
                           INVENTORY_ITEM_ID,
                           SUM(NEW_ORDER_QUANTITY) AS current_supply
                    FROM MSC.MSC_SUPPLIES
                    WHERE ORDER_TYPE IN (1, 2, 3)
                    GROUP BY PLAN_ID, INVENTORY_ITEM_ID
                     ) s ON s.PLAN_ID = ss.PLAN_ID
                         AND s.INVENTORY_ITEM_ID = ss.INVENTORY_ITEM_ID
                WHERE ss.PLAN_ID = :plan_id
                  AND ss.INVENTORY_ITEM_ID = :item_id
            )
            WHERE ROWNUM <= 1
            """,
            binds={"plan_id": plan_id, "item_id": item_id},
            max_rows=1,
        )
        return rows[0] if rows else {}

    def get_pegging_context(
        self,
        item_id: int,
        plan_id: int,
        limit: int = 5,
    ) -> list[dict[str, Any]]:
        return self.execute_query(
            """
            SELECT *
            FROM (
                SELECT fp.PEGGING_ID,
                       fp.DEMAND_ID,
                       fp.TRANSACTION_ID
                           AS supply_transaction_id,
                       s.ORDER_TYPE,
                       s.NEW_ORDER_QUANTITY AS planned_qty,
                       s.FIRM_DATE AS need_by
                FROM MSC.MSC_FULL_PEGGING fp
                JOIN MSC.MSC_SUPPLIES s
                  ON s.PLAN_ID = fp.PLAN_ID
                                 AND s.TRANSACTION_ID = fp.TRANSACTION_ID
                WHERE fp.PLAN_ID = :plan_id
                  AND s.INVENTORY_ITEM_ID = :item_id
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"plan_id": plan_id, "item_id": item_id, "limit": limit},
            max_rows=limit,
        )

    def get_demand_revenue_at_risk(
        self,
        item_id: int,
        plan_id: int,
    ) -> dict[str, Any]:
        """Trace MSC_FULL_PEGGING → MSC_DEMANDS to get actual customer demand value at risk.

        Tutorial Data Flow Step 4: "Revenue at risk? (join to sales orders via pegging)".
        Returns demand_qty (sum of unfulfilled customer demand) and revenue_at_risk
        (demand_qty × demand_unit_price or list_price).  Falls back to {} on any error.
        """
        try:
            rows = self.execute_query(
                """
                SELECT *
                FROM (
                    SELECT fp.DEMAND_ID,
                           d.ORIGINATION_TYPE,
                           d.USING_ASSEMBLY_ITEM_ID,
                           d.DEMAND_DATE,
                           NVL(d.ORIGINATION_QUANTITY, 0)
                               - NVL(d.QUANTITY_COMPLETED, 0) AS unfulfilled_qty,
                           NVL(d.SELLING_PRICE, 0)             AS selling_price,
                           NVL(d.ORIGINATION_QUANTITY, 0)
                               * NVL(d.SELLING_PRICE, 0)       AS demand_revenue
                    FROM MSC.MSC_FULL_PEGGING fp
                    JOIN MSC.MSC_DEMANDS d
                      ON d.PLAN_ID    = fp.PLAN_ID
                     AND d.DEMAND_ID  = fp.DEMAND_ID
                    WHERE fp.PLAN_ID             = :plan_id
                      AND fp.INVENTORY_ITEM_ID   = :item_id
                      AND d.ORIGINATION_TYPE     IN (6, 30)
                      AND NVL(d.ORIGINATION_QUANTITY, 0) > NVL(d.QUANTITY_COMPLETED, 0)
                    ORDER BY demand_revenue DESC
                )
                WHERE ROWNUM <= 20
                """,
                binds={"plan_id": plan_id, "item_id": item_id},
                max_rows=20,
            )
            if not rows:
                return {}
            total_unfulfilled_qty = sum(
                float(r.get("unfulfilled_qty") or 0) for r in rows)
            total_demand_revenue = sum(
                float(r.get("demand_revenue") or 0) for r in rows)
            return {
                "demand_order_count": len(rows),
                "total_unfulfilled_qty": round(total_unfulfilled_qty, 2),
                "total_demand_revenue": round(total_demand_revenue, 2),
                "demand_detail": rows[:5],  # top 5 for evidence
            }
        except Exception:
            return {}

    def get_late_supplier_candidates(
        self,
        days_ahead: int,
        limit: int,
    ) -> list[dict[str, Any]]:
        return self.execute_query(
            """
            SELECT *
            FROM (
                SELECT ph.SEGMENT1 AS po_number,
                       pv.VENDOR_NAME,
                       ph.VENDOR_ID,
                       pl.ITEM_ID,
                       pll.NEED_BY_DATE,
                       TRUNC(SYSDATE)
                           - TRUNC(pll.NEED_BY_DATE)
                           AS days_overdue,
                       pll.QUANTITY
                           - NVL(pll.QUANTITY_RECEIVED, 0)
                           AS qty_outstanding
                FROM {self.tables.po_headers_all} ph
                JOIN {self.tables.po_lines_all} pl ON pl.PO_HEADER_ID = ph.PO_HEADER_ID
                                JOIN APPS.PO_LINE_LOCATIONS_ALL pll
                                    ON pll.PO_LINE_ID = pl.PO_LINE_ID
                JOIN APPS.PO_VENDORS pv ON pv.VENDOR_ID = ph.VENDOR_ID
                WHERE ph.TYPE_LOOKUP_CODE = 'STANDARD'
                  AND ph.AUTHORIZATION_STATUS = 'APPROVED'
                  AND pll.NEED_BY_DATE < TRUNC(SYSDATE) + :days_ahead
                                    AND NVL(pll.QUANTITY_RECEIVED, 0)
                                            < NVL(pll.QUANTITY, 0)
                ORDER BY days_overdue DESC, qty_outstanding DESC
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"days_ahead": days_ahead, "limit": limit},
            max_rows=limit,
        )

    def get_price_anomalies(self, limit: int) -> list[dict[str, Any]]:
        return self.execute_query(
            """
            WITH price_base AS (
                SELECT pl.ITEM_ID,
                       AVG(pl.UNIT_PRICE) AS avg_unit_price,
                       COUNT(*) AS sample_size
                FROM {self.tables.po_lines_all} pl
                                JOIN {self.tables.po_headers_all} ph
                                    ON ph.PO_HEADER_ID = pl.PO_HEADER_ID
                WHERE ph.CREATION_DATE >= ADD_MONTHS(TRUNC(SYSDATE), -12)
                  AND ph.TYPE_LOOKUP_CODE = 'STANDARD'
                GROUP BY pl.ITEM_ID
            )
            SELECT *
            FROM (
                SELECT ph.SEGMENT1 AS po_number,
                       ph.VENDOR_ID,
                       pl.ITEM_ID,
                       pl.UNIT_PRICE,
                       pb.avg_unit_price,
                       ROUND(
                           CASE
                               WHEN pb.avg_unit_price = 0 THEN 0
                               ELSE (
                                   (pl.UNIT_PRICE - pb.avg_unit_price)
                                   / pb.avg_unit_price
                               ) * 100
                           END,
                           2
                       ) AS pct_deviation,
                       pb.sample_size,
                       ph.CREATION_DATE AS po_creation_date
                FROM {self.tables.po_lines_all} pl
                                JOIN {self.tables.po_headers_all} ph
                                    ON ph.PO_HEADER_ID = pl.PO_HEADER_ID
                JOIN price_base pb ON pb.ITEM_ID = pl.ITEM_ID
                WHERE pb.sample_size >= 3
                  AND ph.CREATION_DATE >= ADD_MONTHS(TRUNC(SYSDATE), -3)
                  AND ABS(
                      CASE
                          WHEN pb.avg_unit_price = 0 THEN 0
                          ELSE (
                              (pl.UNIT_PRICE - pb.avg_unit_price)
                              / pb.avg_unit_price
                          ) * 100
                      END
                  ) >= 20
                ORDER BY ABS(
                    CASE
                        WHEN pb.avg_unit_price = 0 THEN 0
                        ELSE (
                            (pl.UNIT_PRICE - pb.avg_unit_price)
                            / pb.avg_unit_price
                        ) * 100
                    END
                ) DESC
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"limit": limit},
            max_rows=limit,
        )

    def get_demand_to_po_gaps(
        self,
        plan_id: int,
        limit: int,
    ) -> list[dict[str, Any]]:
        return self.execute_query(
            """
            SELECT *
            FROM (
                SELECT p.COMPILE_DESIGNATOR AS plan_name,
                       fp.PEGGING_ID,
                       fp.DEMAND_ID,
                       fp.TRANSACTION_ID
                           AS supply_transaction_id,
                       s.INVENTORY_ITEM_ID,
                       s.ORDER_TYPE,
                       s.NEW_ORDER_QUANTITY AS planned_qty,
                       s.FIRM_DATE AS need_by
                FROM MSC.MSC_FULL_PEGGING fp
                JOIN MSC.MSC_PLANS p ON p.PLAN_ID = fp.PLAN_ID
                JOIN MSC.MSC_SUPPLIES s
                  ON s.PLAN_ID = fp.PLAN_ID
                                 AND s.TRANSACTION_ID = fp.TRANSACTION_ID
                WHERE fp.PLAN_ID = :plan_id
                  AND s.ORDER_TYPE = 5
                ORDER BY s.FIRM_DATE NULLS LAST, s.NEW_ORDER_QUANTITY DESC
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"plan_id": plan_id, "limit": limit},
            max_rows=limit,
        )

    def get_spend_summary(self, limit: int) -> list[dict[str, Any]]:
        """Supplier spend summary grouped by supplier + business unit (tutorial spec: Workflow 6)."""
        return self.execute_query(
            """
            SELECT *
            FROM (
                SELECT pv.VENDOR_ID,
                       pv.VENDOR_NAME,
                       ph.ORG_ID                               AS business_unit_id,
                       COUNT(DISTINCT ph.PO_HEADER_ID)         AS po_count,
                       COUNT(DISTINCT pl.ITEM_ID)              AS distinct_items,
                       SUM(pl.QUANTITY * pl.UNIT_PRICE)        AS total_spend,
                       ph.CURRENCY_CODE,
                       MIN(ph.CREATION_DATE)                   AS first_po_date,
                       MAX(ph.CREATION_DATE)                   AS last_po_date
                FROM {self.tables.po_headers_all} ph
                JOIN {self.tables.po_lines_all} pl ON pl.PO_HEADER_ID = ph.PO_HEADER_ID
                JOIN APPS.PO_VENDORS pv ON pv.VENDOR_ID = ph.VENDOR_ID
                WHERE ph.TYPE_LOOKUP_CODE = 'STANDARD'
                  AND ph.AUTHORIZATION_STATUS = 'APPROVED'
                GROUP BY pv.VENDOR_ID, pv.VENDOR_NAME, ph.ORG_ID, ph.CURRENCY_CODE
                ORDER BY SUM(pl.QUANTITY * pl.UNIT_PRICE) DESC NULLS LAST
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"limit": limit},
            max_rows=limit,
        )

    def get_spend_by_time_period(self, limit: int) -> list[dict[str, Any]]:
        """Spend aggregated by quarter and category for time-period trend analysis (tutorial spec: Workflow 6)."""
        try:
            return self.execute_query(
                """
                SELECT *
                FROM (
                    SELECT TO_CHAR(ph.CREATION_DATE, 'YYYY-Q"Q"') AS time_period,
                           TO_CHAR(ph.CREATION_DATE, 'YYYY')      AS year,
                           TO_CHAR(ph.CREATION_DATE, 'Q')         AS quarter,
                           ph.ORG_ID                               AS business_unit_id,
                           COUNT(DISTINCT ph.PO_HEADER_ID)         AS po_count,
                           COUNT(DISTINCT ph.VENDOR_ID)            AS vendor_count,
                           COUNT(DISTINCT pl.ITEM_ID)              AS item_count,
                           SUM(pl.QUANTITY * pl.UNIT_PRICE)        AS total_spend,
                           ph.CURRENCY_CODE
                    FROM {self.tables.po_headers_all} ph
                    JOIN {self.tables.po_lines_all} pl ON pl.PO_HEADER_ID = ph.PO_HEADER_ID
                    WHERE ph.TYPE_LOOKUP_CODE = 'STANDARD'
                    AND ph.AUTHORIZATION_STATUS = 'APPROVED'
                    AND ph.CREATION_DATE >= ADD_MONTHS(TRUNC(SYSDATE), -24)
                    GROUP BY TO_CHAR(ph.CREATION_DATE, 'YYYY-Q"Q"'),
                             TO_CHAR(ph.CREATION_DATE, 'YYYY'),
                             TO_CHAR(ph.CREATION_DATE, 'Q'),
                             ph.ORG_ID, ph.CURRENCY_CODE
                    ORDER BY year DESC, quarter DESC
                )
                WHERE ROWNUM <= :limit
                """,
                binds={"limit": limit},
                max_rows=limit,
            )
        except Exception:
            return []

    def get_maverick_spend(self, limit: int) -> list[dict[str, Any]]:
        """Detect purchases made outside of active blanket agreements (maverick spend)."""
        try:
            return self.execute_query(
                """
                SELECT *
                FROM (
                    SELECT ph.SEGMENT1 AS po_number,
                           pv.VENDOR_NAME,
                           pl.ITEM_ID AS inventory_item_id,
                           pl.UNIT_PRICE,
                           pl.QUANTITY,
                           ROUND(pl.QUANTITY * pl.UNIT_PRICE, 2) AS line_spend,
                           ph.CREATION_DATE
                    FROM {self.tables.po_headers_all} ph
                    JOIN {self.tables.po_lines_all} pl ON pl.PO_HEADER_ID = ph.PO_HEADER_ID
                    JOIN APPS.PO_VENDORS pv ON pv.VENDOR_ID = ph.VENDOR_ID
                    WHERE ph.TYPE_LOOKUP_CODE = 'STANDARD'
                    AND ph.AUTHORIZATION_STATUS = 'APPROVED'
                    AND ph.CREATION_DATE >= TRUNC(SYSDATE) - 365
                    AND NOT EXISTS (
                        SELECT 1 FROM {self.tables.po_headers_all} pa
                        WHERE pa.TYPE_LOOKUP_CODE IN ('BLANKET', 'CONTRACT')
                        AND pa.VENDOR_ID = ph.VENDOR_ID
                        AND pa.AUTHORIZATION_STATUS = 'APPROVED'
                    )
                    ORDER BY pl.QUANTITY * pl.UNIT_PRICE DESC NULLS LAST
                )
                WHERE ROWNUM <= :limit
                """,
                binds={"limit": limit},
                max_rows=limit,
            )
        except Exception:
            return []

    def get_single_source_items(self, limit: int) -> list[dict[str, Any]]:
        """Identify items sourced from only one supplier (supply chain risk)."""
        try:
            return self.execute_query(
                """
                SELECT *
                FROM (
                    SELECT pl.ITEM_ID AS inventory_item_id,
                           COUNT(DISTINCT ph.VENDOR_ID) AS supplier_count,
                           SUM(pl.QUANTITY * pl.UNIT_PRICE) AS total_spend,
                           MAX(pv.VENDOR_NAME) AS sole_supplier_name
                    FROM {self.tables.po_headers_all} ph
                    JOIN {self.tables.po_lines_all} pl ON pl.PO_HEADER_ID = ph.PO_HEADER_ID
                    JOIN APPS.PO_VENDORS pv ON pv.VENDOR_ID = ph.VENDOR_ID
                    WHERE ph.TYPE_LOOKUP_CODE = 'STANDARD'
                    AND ph.AUTHORIZATION_STATUS = 'APPROVED'
                    AND ph.CREATION_DATE >= TRUNC(SYSDATE) - 365
                    AND pl.ITEM_ID IS NOT NULL
                    GROUP BY pl.ITEM_ID
                    HAVING COUNT(DISTINCT ph.VENDOR_ID) = 1
                    ORDER BY SUM(pl.QUANTITY * pl.UNIT_PRICE) DESC NULLS LAST
                )
                WHERE ROWNUM <= :limit
                """,
                binds={"limit": limit},
                max_rows=limit,
            )
        except Exception:
            return []

    def get_consolidation_opportunities(self, limit: int) -> list[dict[str, Any]]:
        """Find items purchased from multiple suppliers where consolidation could save cost."""
        try:
            return self.execute_query(
                """
                SELECT *
                FROM (
                    SELECT pl.ITEM_ID AS inventory_item_id,
                           COUNT(DISTINCT ph.VENDOR_ID) AS supplier_count,
                           COUNT(DISTINCT ph.PO_HEADER_ID) AS po_count,
                           ROUND(MIN(pl.UNIT_PRICE), 4) AS min_price,
                           ROUND(MAX(pl.UNIT_PRICE), 4) AS max_price,
                           ROUND(AVG(pl.UNIT_PRICE), 4) AS avg_price,
                           ROUND(MAX(pl.UNIT_PRICE) - MIN(pl.UNIT_PRICE), 4) AS price_spread,
                           SUM(pl.QUANTITY * pl.UNIT_PRICE) AS total_spend
                    FROM {self.tables.po_headers_all} ph
                    JOIN {self.tables.po_lines_all} pl ON pl.PO_HEADER_ID = ph.PO_HEADER_ID
                    WHERE ph.TYPE_LOOKUP_CODE = 'STANDARD'
                    AND ph.AUTHORIZATION_STATUS = 'APPROVED'
                    AND ph.CREATION_DATE >= TRUNC(SYSDATE) - 365
                    AND pl.ITEM_ID IS NOT NULL
                    AND pl.UNIT_PRICE > 0
                    GROUP BY pl.ITEM_ID
                    HAVING COUNT(DISTINCT ph.VENDOR_ID) > 1
                    AND (MAX(pl.UNIT_PRICE) - MIN(pl.UNIT_PRICE)) / MIN(pl.UNIT_PRICE) > 0.10
                    ORDER BY SUM(pl.QUANTITY * pl.UNIT_PRICE) DESC NULLS LAST
                )
                WHERE ROWNUM <= :limit
                """,
                binds={"limit": limit},
                max_rows=limit,
            )
        except Exception:
            return []

    def list_organization_ids(self) -> dict[str, Any]:
        """Get all available organization IDs from configured table.

        Returns a dict with total count and list of organizations.
        """
        rows = self.execute_query(
            f"""
            SELECT DISTINCT
                   mp.ORGANIZATION_ID,
                   mp.ORGANIZATION_CODE
            FROM {self.tables.mtp_parameters} mp
            ORDER BY mp.ORGANIZATION_ID
            """,
            max_rows=1000,
        )
        return {
            "total_count": len(rows),
            "organizations": rows,
        }


class TutorialProcurementAgent:
    EXCEPTION_LABELS = {
        1: "Items with No Activity",
        2: "Late Replenishment Orders",
        3: "Items Below Safety Stock",
        4: "Excess Inventory",
        5: "Expired Lots",
        6: "Past Due Purchase Orders",
        7: "Cancelled Supply",
        8: "Short Shipments",
        9: "Supplier Capacity Breach",
        10: "Lead Time Breach",
        11: "Orders at Risk",
        12: "Demand Past Due",
        13: "Demand Not Satisfied",
        14: "Late Demand",
        21: "Resource Overloaded",
        22: "Resource Underloaded",
        23: "Resource Unavailable",
    }

    def __init__(
        self,
        gateway: OracleReadOnlyGateway,
        autonomy_level: int = 1,
        organization_id: int | None = None,
    ) -> None:
        self.gateway = gateway
        self.autonomy_level = autonomy_level
        self.organization_id = organization_id
        self.actions: list[ActionRecord] = []
        self.decision_engine = ProcurementDecisionEngine()
        self._draft_po_seq = 0  # sequence for human-readable PO numbers

    def reset_actions(self) -> None:
        self.actions = []

    def run(
        self,
        workflow: str,
        plan_id: int | None,
        limit: int,
    ) -> dict[str, Any]:
        summary = self.gateway.get_exception_summary(limit=5)
        selected_plan_id = plan_id or self._select_plan_id(summary)
        result: dict[str, Any] = {
            "autonomy_level": self.autonomy_level,
            "organization_id": self.organization_id,
            "selected_plan_id": selected_plan_id,
            "selected_plan_name": self._select_plan_name(
                summary,
                selected_plan_id,
            ),
            "workflows": {},
        }

        if workflow in {"exception-triage", "all"}:
            result["workflows"]["exception-triage"] = (
                self.run_exception_triage(selected_plan_id, limit)
            )
        if workflow in {"late-supplier", "all"}:
            result["workflows"]["late-supplier"] = (
                self.run_late_supplier_detection(limit)
            )
        if workflow in {"safety-stock", "all"}:
            result["workflows"]["safety-stock"] = (
                self.run_safety_stock_alerts(selected_plan_id, limit)
            )
        if workflow in {"price-anomaly", "all"}:
            result["workflows"]["price-anomaly"] = (
                self.run_price_anomaly_detection(limit)
            )
        if workflow in {"demand-to-po", "all"}:
            result["workflows"]["demand-to-po"] = (
                self.run_demand_to_po_tracing(selected_plan_id, limit)
            )
        if workflow in {"spend-analytics", "all"}:
            result["workflows"]["spend-analytics"] = (
                self.run_spend_analytics(limit)
            )

        action_counter = Counter(action.action for action in self.actions)
        result["action_summary"] = dict(action_counter)
        result["actions"] = [asdict(action) for action in self.actions]
        return result

    def _select_plan_id(self, summary: dict[str, Any]) -> int:
        top_plans = summary.get("top_plans", [])
        if not top_plans:
            raise RuntimeError("No MSC plans with exceptions were returned.")
        return int(top_plans[0]["plan_id"])

    def _select_plan_name(self, summary: dict[str, Any], plan_id: int) -> str:
        for plan in summary.get("top_plans", []):
            if int(plan["plan_id"]) == int(plan_id):
                return str(plan.get("plan_name") or plan_id)
        return str(plan_id)

    def _exception_label(self, exception_type: int) -> str:
        return self.EXCEPTION_LABELS.get(
            exception_type,
            f"Exception Type {exception_type}",
        )

    def _record_action(
        self,
        workflow: str,
        action: str,
        priority: str,
        requires_human_approval: bool,
        summary: str,
        rationale: str,
        evidence: dict[str, Any],
        payload: dict[str, Any],
        confidence: float | None = None,
        assigned_to: str | None = None,
    ) -> dict[str, Any]:
        if self.autonomy_level == 0 and action != "report_insight":
            payload = {
                "recommended_action": action,
                "deferred_payload": payload,
            }
            action = "report_only"
            requires_human_approval = False

        entry = ActionRecord(
            workflow=workflow,
            action=action,
            priority=priority,
            requires_human_approval=requires_human_approval,
            summary=summary,
            rationale=rationale,
            evidence=evidence,
            payload=payload,
            confidence=confidence,
            assigned_to=assigned_to,
        )
        self.actions.append(entry)
        return asdict(entry)

    def _get_org_id(self) -> int:
        """Get default organization/operating unit from session context."""
        if self.organization_id is not None:
            try:
                result = self.gateway.execute_query(
                    """
                    SELECT ORGANIZATION_ID
                    FROM MTL_PARAMETERS
                    WHERE ORGANIZATION_ID = :organization_id
                    """,
                    binds={"organization_id": self.organization_id},
                    max_rows=1,
                )
                if result:
                    return int(result[0]["organization_id"])
                raise ValueError(
                    f"Organization ID {self.organization_id} was not found in MTL_PARAMETERS"
                )
            except ValueError:
                raise
            except Exception as exc:
                raise RuntimeError(
                    f"Failed to validate organization ID {self.organization_id}"
                ) from exc

        try:
            result = self.gateway.execute_query(
                "SELECT ORGANIZATION_ID FROM MTL_PARAMETERS WHERE ROWNUM = 1",
                max_rows=1
            )
            return int(result[0]["organization_id"]) if result else None
        except Exception:
            return None

    def _get_ship_to_location(self, item_id: int | None = None) -> dict[str, Any]:
        """Query default ship-to location from HR_LOCATIONS."""
        try:
            sql = """
                SELECT LOCATION_ID, LOCATION_CODE, LOCATION_NAME
                FROM HR.HR_LOCATIONS
                WHERE INACTIVE_DATE IS NULL
                ORDER BY LOCATION_ID
            """
            result = self.gateway.execute_query(sql, max_rows=1)
            return result[0] if result else {}
        except Exception:
            return {}

    def _get_supplier_terms(self, vendor_id: int, vendor_site_id: int | None) -> dict[str, Any]:
        """Query supplier site details: payment terms, FOB, freight."""
        try:
            sql = """
                SELECT
                    pvs.VENDOR_ID,
                    pvs.VENDOR_SITE_ID,
                    pvs.PAYMENT_TERMS_ID,
                    pvs.FOB_LOOKUP_CODE,
                    pvs.FREIGHT_TERMS_LOOKUP_CODE,
                    pvs.LEAD_TIME_DAYS,
                    pt.TERM_NAME
                FROM PO.PO_VENDOR_SITES_ALL pvs
                LEFT JOIN AP.AP_TERMS pt ON pt.TERM_ID = pvs.PAYMENT_TERMS_ID
                WHERE pvs.VENDOR_ID = :vendor_id
                AND (pvs.VENDOR_SITE_ID = :vendor_site_id OR :vendor_site_id IS NULL)
                AND pvs.INACTIVE_DATE IS NULL
                ORDER BY pvs.LEAD_TIME_DAYS ASC
            """
            result = self.gateway.execute_query(
                sql,
                binds={"vendor_id": vendor_id,
                       "vendor_site_id": vendor_site_id},
                max_rows=1
            )
            return result[0] if result else {}
        except Exception:
            return {}

    def _get_item_details(self, item_id: int) -> dict[str, Any]:
        """Query item master: description, UOM, price."""
        try:
            sql = """
                SELECT
                    INVENTORY_ITEM_ID,
                    SEGMENT1 AS ITEM_NUMBER,
                    DESCRIPTION,
                    PRIMARY_UOM_CODE,
                    LIST_PRICE_PER_UNIT,
                    PURCHASING_ENABLED_FLAG,
                    INVENTORY_ITEM_FLAG
                FROM MTL.MTL_SYSTEM_ITEMS_B
                WHERE INVENTORY_ITEM_ID = :item_id
                AND INACTIVE_DATE IS NULL
            """
            result = self.gateway.execute_query(
                sql,
                binds={"item_id": item_id},
                max_rows=1
            )
            return result[0] if result else {}
        except Exception:
            return {}

    def _lookup_unit_price(self, item_id: int, vendor_id: int | None = None) -> float | None:
        """Look up unit price from supplier history or contract."""
        try:
            if vendor_id:
                # Try to get from supplier contract/agreement
                sql = """
                    SELECT AVG(pl.UNIT_PRICE) AS avg_unit_price
                    FROM {self.tables.po_lines_all} pl
                    JOIN PO.PO_HEADERS_ALL ph ON ph.PO_HEADER_ID = pl.PO_HEADER_ID
                    WHERE pl.INVENTORY_ITEM_ID = :item_id
                    AND ph.VENDOR_ID = :vendor_id
                    AND ph.AUTHORIZATION_STATUS IN ('APPROVED', 'CLOSED')
                    AND ph.CREATION_DATE >= TRUNC(SYSDATE) - 365
                """
                result = self.gateway.execute_query(
                    sql,
                    binds={"item_id": item_id, "vendor_id": vendor_id},
                    max_rows=1
                )
                if result and result[0].get("avg_unit_price"):
                    return float(result[0]["avg_unit_price"])
        except Exception:
            pass

        # Fallback to item list price
        try:
            item = self._get_item_details(item_id)
            return item.get("list_price_per_unit")
        except Exception:
            return None

    def _get_charge_account(self, org_id: int | None = None) -> dict[str, Any]:
        """Query GL charge account (CCID) for inventory distribution."""
        try:
            sql = """
                SELECT
                    CHART_OF_ACCOUNTS_ID,
                    SEGMENT1 || '.' || SEGMENT2 || '.' || SEGMENT3 AS ACCOUNT_CODE,
                    DESCRIPTION
                FROM GL.GL_CODE_COMBINATIONS
                WHERE ACCOUNT_TYPE = 'A'  -- Asset accounts
                AND ENABLED_FLAG = 'Y'
                AND SEGMENT1 = '1010'  -- Example inventory account
                AND ROWNUM = 1
            """
            result = self.gateway.execute_query(sql, max_rows=1)
            return result[0] if result else {}
        except Exception:
            return {}

    def _validate_r12_fields(self, payload: dict[str, Any]) -> tuple[bool, str]:
        """Validate all required R12 fields before inserting."""
        header = payload.get("header", {})
        line = payload.get("line", {})

        required_fields = {
            "header": ["segment1", "type_lookup_code", "vendor_id", "org_id"],
            "line": ["item_id", "quantity"],
            "distribution": ["charge_account"]
        }

        for section, fields in required_fields.items():
            section_data = payload.get(section, {})
            for field in fields:
                if not section_data.get(field):
                    return False, f"Missing required field: {section}.{field}"

        return True, "OK"

    def validate_procurement_rules(
        self,
        item_id: int,
        vendor_id: int | None,
        quantity: float,
        unit_price: float,
        need_by_date_str: str | None,
    ) -> dict[str, Any]:
        """Run Phase 5 procurement validation rules before creating a PO.

        Rules:
        1. Supplier active status - reject if inactive
        2. Item purchasable flag - reject if not purchasable
        3. Lead-time feasibility - warn if lead time exceeds need_by_date
        4. MOQ validation - warn if quantity < minimum order quantity
        5. Financial threshold - warn if PO value > $100K (needs approval)

        Returns:
            {
                "can_create_po": bool,
                "warnings": list[str],
                "blocks": list[str],
                "supplier_status": str,
                "item_status": str,
                "lead_time_days": int | None,
                "moq": float | None,
                "estimated_po_value": float,
            }
        """
        blocks = []
        warnings = []
        supplier_status = "UNKNOWN"
        item_status = "UNKNOWN"
        lead_time_days = None
        moq = None

        # Rule 1: Supplier active status
        if vendor_id:
            try:
                result = self.gateway.execute_query(
                    """
                    SELECT VENDOR_NAME,
                           END_DATE_ACTIVE,
                           HOLD_FLAG,
                           VENDOR_TYPE_LOOKUP_CODE
                    FROM APPS.PO_VENDORS
                    WHERE VENDOR_ID = :vendor_id
                    """,
                    binds={"vendor_id": vendor_id},
                    max_rows=1,
                )
                if result:
                    supplier = result[0]
                    end_date = supplier.get("end_date_active")
                    hold_flag = supplier.get("hold_flag")
                    if hold_flag == "Y":
                        blocks.append(
                            f"Supplier {supplier.get('vendor_name')} is ON HOLD - cannot issue PO")
                        supplier_status = "ON_HOLD"
                    elif end_date and end_date < datetime.now():
                        blocks.append(
                            f"Supplier {supplier.get('vendor_name')} is INACTIVE (end date: {end_date})")
                        supplier_status = "INACTIVE"
                    else:
                        supplier_status = "ACTIVE"
                else:
                    warnings.append(
                        f"Vendor ID {vendor_id} not found in PO_VENDORS")
                    supplier_status = "NOT_FOUND"
            except Exception:
                warnings.append(
                    "Could not validate supplier status (PO_VENDORS query failed)")

        # Rule 2: Item purchasable flag
        if item_id:
            try:
                result = self.gateway.execute_query(
                    """
                    SELECT SEGMENT1, PURCHASING_ENABLED_FLAG,
                           END_DATE_ACTIVE, INVENTORY_ITEM_ID
                    FROM INV.MTL_SYSTEM_ITEMS_B
                    WHERE INVENTORY_ITEM_ID = :item_id
                    AND ROWNUM = 1
                    """,
                    binds={"item_id": item_id},
                    max_rows=1,
                )
                if result:
                    item = result[0]
                    if item.get("purchasing_enabled_flag") != "Y":
                        blocks.append(
                            f"Item {item.get('segment1', item_id)} is not purchasing-enabled "
                            f"(PURCHASING_ENABLED_FLAG = {item.get('purchasing_enabled_flag')})"
                        )
                        item_status = "NOT_PURCHASABLE"
                    elif item.get("end_date_active") and item.get("end_date_active") < datetime.now():
                        blocks.append(
                            f"Item {item.get('segment1', item_id)} is INACTIVE (past END_DATE_ACTIVE)")
                        item_status = "INACTIVE"
                    else:
                        item_status = "PURCHASABLE"
                else:
                    warnings.append(
                        f"Item {item_id} not found in MTL_SYSTEM_ITEMS_B")
                    item_status = "NOT_FOUND"
            except Exception:
                warnings.append(
                    "Could not validate item status (MTL_SYSTEM_ITEMS_B query failed)")

        # Rule 3: Lead-time feasibility
        if vendor_id and need_by_date_str:
            try:
                result = self.gateway.execute_query(
                    """
                    SELECT LEAD_TIME_DAYS
                    FROM APPS.PO_VENDOR_SITES_ALL
                    WHERE VENDOR_ID = :vendor_id
                    AND INACTIVE_DATE IS NULL
                    AND ROWNUM = 1
                    """,
                    binds={"vendor_id": vendor_id},
                    max_rows=1,
                )
                if result and result[0].get("lead_time_days"):
                    lead_time_days = int(result[0]["lead_time_days"])
                    need_by = datetime.fromisoformat(
                        str(need_by_date_str)) if need_by_date_str else None
                    if need_by:
                        days_until_needed = (need_by - datetime.now()).days
                        if lead_time_days > days_until_needed:
                            warnings.append(
                                f"Lead time ({lead_time_days} days) exceeds days until needed "
                                f"({days_until_needed} days). Consider alternate supplier or expedite."
                            )
            except Exception:
                pass  # Lead time check is a warning only

        # Rule 4: MOQ validation
        if vendor_id and item_id:
            try:
                result = self.gateway.execute_query(
                    """
                    SELECT MIN_ORDER_QTY
                    FROM MSC.MSC_ITEM_SUPPLIERS
                    WHERE INVENTORY_ITEM_ID = :item_id
                    AND SUPPLIER_ID = :vendor_id
                    AND ROWNUM = 1
                    """,
                    binds={"item_id": item_id, "vendor_id": vendor_id},
                    max_rows=1,
                )
                if result:
                    moq = float(result[0].get("min_order_qty") or 0)
                    if moq and quantity < moq:
                        warnings.append(
                            f"Quantity {quantity} is below minimum order quantity "
                            f"({moq}) for this supplier-item combination."
                        )
            except Exception:
                pass  # MOQ check is a warning only

        # Rule 5: Financial threshold check
        estimated_po_value = quantity * (unit_price or 0)
        if estimated_po_value > 100000:
            warnings.append(
                f"High-value PO: estimated value ${estimated_po_value:,.2f} exceeds $100K. "
                "Procurement manager approval required before issuing."
            )

        return {
            "can_create_po": len(blocks) == 0,
            "warnings": warnings,
            "blocks": blocks,
            "supplier_status": supplier_status,
            "item_status": item_status,
            "lead_time_days": lead_time_days,
            "moq": moq,
            "estimated_po_value": round(estimated_po_value, 2),
        }

    def _create_draft_po_payload(
        self,
        item_id: int,
        quantity: float,
        need_by_date: Any,
        reason: str,
        supplier_options: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Create complete Oracle R12 EBS draft PO payload with all required fields."""
        self._draft_po_seq += 1
        date_part = datetime.now().strftime("%Y%m%d")
        po_number = f"AGNT-{date_part}-{self._draft_po_seq:04d}"
        vendor = supplier_options[0] if supplier_options else {}
        quantity_value = float(quantity or 0)

        # Get Oracle R12 context data
        org_id = self._get_org_id()
        item_details = self._get_item_details(item_id)
        supplier_terms = self._get_supplier_terms(
            vendor.get("vendor_id"),
            vendor.get("vendor_site_id")
        )
        ship_to_location = self._get_ship_to_location(item_id)
        unit_price = self._lookup_unit_price(item_id, vendor.get("vendor_id"))
        charge_account = self._get_charge_account(org_id)

        line_amount = quantity_value * (unit_price or 0)

        return {
            "po_number": po_number,
            "status": "DRAFT - PENDING BUYER APPROVAL",
            "header": {
                # Core PO identification
                "segment1": po_number,
                "po_number": po_number,
                "type_lookup_code": "STANDARD",
                "authorization_status": "INCOMPLETE",

                # Vendor info
                "vendor_id": vendor.get("vendor_id"),
                "vendor_name": vendor.get("vendor_name"),
                "vendor_site_id": vendor.get("vendor_site_id"),
                "vendor_site_code": vendor.get("vendor_site_code"),

                # Organization context (CRITICAL FOR R12)
                "org_id": org_id,
                "operating_unit_id": org_id,

                # Terms and conditions
                "currency_code": "USD",
                "payment_terms_id": supplier_terms.get("payment_terms_id"),
                "fob_lookup_code": supplier_terms.get("fob_lookup_code", "SHIPPER"),
                "freight_terms_lookup_code": supplier_terms.get("freight_terms_lookup_code", "COLLECT"),

                # Delivery location
                "ship_to_location_id": ship_to_location.get("location_id"),
                "ship_to_location_code": ship_to_location.get("location_code"),
                "bill_to_location_id": ship_to_location.get("location_id"),

                # PO flags
                "acceptance_required_flag": "N",
                "receiving_required_flag": "Y",
                "invoice_match_option_lookup_code": "3-way",  # 3-way match

                # Comments and metadata
                "comments": reason,
                "created_by_module": "TUTORIAL_AGENT",
                "creation_date": datetime.now().isoformat(),
            },
            "line": {
                # Item information
                "item_id": item_id,
                "inventory_item_id": item_id,
                "item_number": item_details.get("segment1"),
                "item_description": item_details.get("description"),
                "unit_of_measure": item_details.get("primary_uom_code", "EA"),
                "unit_meas_lookup_code": item_details.get("primary_uom_code", "EA"),

                # Quantity and pricing
                "quantity": quantity_value,
                "unit_price": unit_price,
                "line_amount": line_amount,

                # Line number (would be 1 for single-line PO)
                "line_num": 1,

                # Item type flags
                "item_type_code": "GOODS",
                "need_by_date": need_by_date,
                "receipt_required_flag": "Y",
            },
            "shipment": {
                # Schedule/delivery information
                "need_by_date": need_by_date,
                "quantity_ordered": quantity_value,

                # Delivery location
                "ship_to_location_id": ship_to_location.get("location_id"),
                "location_id": ship_to_location.get("location_id"),

                # Receipt flags
                "receipt_required_flag": "Y",
                "inspection_required_flag": "N",

                # Matching requirements
                "receipt_days_exception_code": "NONE",
            },
            "distribution": {
                # Inventory destination
                "destination_type_code": "INVENTORY",
                "quantity_ordered": quantity_value,

                # GL account (CRITICAL - where charge goes)
                "charge_account": charge_account.get("account_code") or "1010-1000-1000",
                "code_combination_id": charge_account.get("chart_of_accounts_id"),

                # Amount
                "amount_ordered": line_amount,

                # Project accounting (if applicable)
                "project_id": None,
                "task_id": None,
                "expenditure_type": None,

                # Distribution metadata
                "distribution_type": "STANDARD",
            },

            # Run Phase 5 procurement validation rules
            "validation_status": {
                "org_id_found": org_id is not None,
                "charge_account_found": bool(charge_account),
                "supplier_details_found": bool(supplier_terms),
                "item_details_found": bool(item_details),
                **self.validate_procurement_rules(
                    item_id=item_id,
                    vendor_id=vendor.get("vendor_id"),
                    quantity=quantity_value,
                    unit_price=unit_price or 0,
                    need_by_date_str=str(
                        need_by_date) if need_by_date else None,
                ),
            }
        }

    def _action_or_recommendation(
        self,
        workflow: str,
        priority: str,
        summary: str,
        rationale: str,
        evidence: dict[str, Any],
        payload: dict[str, Any],
        preferred_action: str,
        confidence: float | None = None,
    ) -> dict[str, Any]:
        if preferred_action == "create_draft_po" and self.autonomy_level < 2:
            return self._record_action(
                workflow=workflow,
                action="recommend_draft_po",
                priority=priority,
                requires_human_approval=True,
                summary=summary,
                rationale=rationale,
                evidence=evidence,
                payload=payload,
                confidence=confidence,
            )
        return self._record_action(
            workflow=workflow,
            action=preferred_action,
            priority=priority,
            requires_human_approval=preferred_action != "auto_resolve",
            summary=summary,
            rationale=rationale,
            evidence=evidence,
            payload=payload,
            confidence=confidence,
        )

    def create_draft_po_action(
        self,
        workflow: str,
        item_id: int,
        quantity: float,
        need_by_date: Any,
        reason: str,
        priority: str,
        evidence: dict[str, Any],
        confidence: float | None = None,
    ) -> dict[str, Any]:
        suppliers = self.gateway.get_supplier_options(item_id)
        payload = self._create_draft_po_payload(
            item_id=item_id,
            quantity=quantity,
            need_by_date=need_by_date,
            reason=reason,
            supplier_options=suppliers,
        )
        summary = (
            f"Draft PO prepared for item {item_id} with "
            f"quantity {quantity:,.2f}"
        )
        return self._action_or_recommendation(
            workflow=workflow,
            priority=priority,
            summary=summary,
            rationale=reason,
            evidence=evidence,
            payload=payload,
            preferred_action="create_draft_po",
            confidence=confidence,
        )

    def notify_buyer_action(
        self,
        workflow: str,
        priority: str,
        summary: str,
        rationale: str,
        evidence: dict[str, Any],
        revenue_at_risk: float | None = None,
        confidence: float | None = None,
        assigned_to: str = "Senior Buyer",
    ) -> dict[str, Any]:
        payload = {
            "notify": "buyer",
            "revenue_at_risk": revenue_at_risk,
            "assigned_to": assigned_to,
        }
        return self._record_action(
            workflow=workflow,
            action="notify_buyer",
            priority=priority,
            requires_human_approval=True,
            summary=summary,
            rationale=rationale,
            evidence=evidence,
            payload=payload,
            confidence=confidence,
            assigned_to=assigned_to,
        )

    def add_to_worklist_action(
        self,
        workflow: str,
        priority: str,
        summary: str,
        rationale: str,
        evidence: dict[str, Any],
        bucket: str,
        confidence: float | None = None,
    ) -> dict[str, Any]:
        return self._record_action(
            workflow=workflow,
            action="add_to_worklist",
            priority=priority,
            requires_human_approval=True,
            summary=summary,
            rationale=rationale,
            evidence=evidence,
            payload={"worklist_bucket": bucket},
            confidence=confidence,
        )

    def auto_resolve_action(
        self,
        workflow: str,
        priority: str,
        summary: str,
        rationale: str,
        evidence: dict[str, Any],
        resolution: str,
        confidence: float | None = None,
    ) -> dict[str, Any]:
        return self._record_action(
            workflow=workflow,
            action="auto_resolve",
            priority=priority,
            requires_human_approval=False,
            summary=summary,
            rationale=rationale,
            evidence=evidence,
            payload={"resolution": resolution},
            confidence=confidence,
        )

    def report_insight_action(
        self,
        workflow: str,
        summary: str,
        rationale: str,
        evidence: dict[str, Any],
        next_step: str,
        confidence: float | None = None,
    ) -> dict[str, Any]:
        return self._record_action(
            workflow=workflow,
            action="report_insight",
            priority="P3-MEDIUM",
            requires_human_approval=False,
            summary=summary,
            rationale=rationale,
            evidence=evidence,
            payload={"recommended_next_step": next_step},
            confidence=confidence,
        )

    def run_exception_triage(self, plan_id: int, limit: int) -> dict[str, Any]:
        exception_types = self.gateway.get_exception_types(plan_id)
        supported_types = {1, 2, 3, 4, 5, 6, 11, 13}
        filtered_types = [
            row
            for row in exception_types
            if int(row["exception_type"]) in supported_types
        ]
        top_types = filtered_types[:3] or exception_types[:3]
        decisions: list[dict[str, Any]] = []

        for exc_type_row in top_types:
            exception_type = int(exc_type_row["exception_type"])
            details = self.gateway.get_exception_details(
                plan_id,
                exception_type,
                limit=max(3, limit),
            )
            for record in details:
                decisions.append(
                    self._triage_exception(plan_id, exception_type, record)
                )

        return {
            "plan_id": plan_id,
            "exception_types": top_types,
            "decisions_made": len(decisions),
            "sample_decisions": decisions[: min(10, len(decisions))],
        }

    def _triage_exception(
        self,
        plan_id: int,
        exception_type: int,
        record: dict[str, Any],
    ) -> dict[str, Any]:
        item_id = int(record.get("inventory_item_id") or 0)
        quantity = float(record.get("quantity") or 0)
        item_context = (
            self.gateway.get_item_context(item_id) if item_id else {}
        )
        supplier_options = (
            self.gateway.get_supplier_options(item_id) if item_id else []
        )
        open_po_coverage = (
            self.gateway.get_open_po_coverage(item_id) if item_id else []
        )
        safety_stock = (
            self.gateway.get_safety_stock_context(item_id, plan_id)
            if item_id
            else {}
        )
        pegging = (
            self.gateway.get_pegging_context(item_id, plan_id)
            if item_id
            else []
        )

        unit_price = float(item_context.get("list_price_per_unit") or 0)

        # Use decision engine to score exception priority
        engine_score = self.decision_engine.score_exception_priority(
            exception_type=exception_type,
            quantity=quantity,
            days_overdue=0,
            unit_price=unit_price,
            pegging_count=len(pegging),
        )

        evidence = {
            "plan_id": plan_id,
            "exception_id": record.get("exception_detail_id"),
            "exception_type": exception_type,
            "exception_label": self._exception_label(exception_type),
            "item_id": item_id,
            "quantity": quantity,
            "item_context": item_context,
            "supplier_options": supplier_options[:2],
            "open_po_coverage": open_po_coverage[:2],
            "safety_stock": safety_stock,
            "pegging_count": len(pegging),
            "decision_engine_score": engine_score,
        }

        summary = (
            f"{self._exception_label(exception_type)} for item {item_id} "
            f"with quantity signal {quantity:,.2f}"
        )

        if exception_type == 2 and quantity >= 100:
            payload = self._create_draft_po_payload(
                item_id=item_id,
                quantity=quantity,
                need_by_date=record.get("date1"),
                reason=(
                    "Late replenishment detected by tutorial "
                    "exception triage workflow."
                ),
                supplier_options=supplier_options,
            )
            return self._action_or_recommendation(
                workflow="exception-triage",
                priority="P1-CRITICAL",
                summary=summary,
                rationale=(
                    "Large replenishment shortage detected; the tutorial "
                    "treats this as high-risk and suitable for urgent "
                    "follow-up procurement."
                ),
                evidence=evidence,
                payload=payload,
                preferred_action="create_draft_po",
            )

        if (
            exception_type == 3
            and max(quantity, float(safety_stock.get("shortage") or 0)) >= 50
        ):
            shortage = max(quantity, float(safety_stock.get("shortage") or 0))
            payload = self._create_draft_po_payload(
                item_id=item_id,
                quantity=shortage,
                need_by_date=record.get("date1"),
                reason="Safety stock breach detected by tutorial workflow.",
                supplier_options=supplier_options,
            )
            return self._action_or_recommendation(
                workflow="exception-triage",
                priority="P2-HIGH",
                summary=summary,
                rationale=(
                    "Current supply is below safety stock target; the "
                    "tutorial says to create urgent replenishment routed "
                    "to an approved supplier."
                ),
                evidence=evidence,
                payload=payload,
                preferred_action="create_draft_po",
            )

        if exception_type == 6 and quantity > 0:
            payload = self._create_draft_po_payload(
                item_id=item_id,
                quantity=quantity,
                need_by_date=record.get("date1"),
                reason=(
                    "Past due purchase order exception triggered follow-up "
                    "procurement review."
                ),
                supplier_options=supplier_options,
            )
            return self._action_or_recommendation(
                workflow="exception-triage",
                priority="P2-HIGH",
                summary=summary,
                rationale=(
                    "The tutorial identifies past due POs as active "
                    "procurement issues that should trigger follow-up "
                    "ordering or escalation."
                ),
                evidence=evidence,
                payload=payload,
                preferred_action="create_draft_po",
            )

        if exception_type in {11, 13}:
            # Tutorial Data Flow Step 4: trace pegging → actual customer demand exposure
            demand_risk = (
                self.gateway.get_demand_revenue_at_risk(item_id, plan_id)
                if item_id
                else {}
            )
            if demand_risk.get("total_demand_revenue", 0) > 0:
                # Pegging trace succeeded — use actual customer order value
                revenue_at_risk = demand_risk["total_demand_revenue"]
                revenue_source = "pegging_demand_trace"
            else:
                # Fallback: quantity × list price (coarse estimate)
                revenue_at_risk = round(
                    quantity *
                    float(item_context.get("list_price_per_unit") or 0),
                    2,
                )
                revenue_source = "list_price_estimate"
            payload = {
                "notify": "buyer",
                "assigned_to": "Senior Buyer",
                "revenue_at_risk": revenue_at_risk,
                "revenue_calculation_source": revenue_source,
                "demand_orders_affected": demand_risk.get("demand_order_count", 0),
                "unfulfilled_demand_qty": demand_risk.get("total_unfulfilled_qty", 0),
            }
            return self._record_action(
                workflow="exception-triage",
                action="notify_buyer",
                priority="P1-CRITICAL" if exception_type == 11 else "P2-HIGH",
                requires_human_approval=True,
                summary=summary,
                rationale=(
                    "Demand-facing exceptions require human judgement "
                    "because customer fulfillment may be at risk."
                ),
                evidence=evidence,
                payload=payload,
            )

        if exception_type == 4 and quantity <= 500:
            return self._record_action(
                workflow="exception-triage",
                action="auto_resolve",
                priority="P3-MEDIUM",
                requires_human_approval=False,
                summary=summary,
                rationale=(
                    "Small excess inventory is low-risk and can be moved "
                    "to the worklist or auto-resolved under the tutorial "
                    "model."
                ),
                evidence=evidence,
                payload={
                    "resolution": (
                        "Monitor only; no replenishment action required."
                    )
                },
            )

        if exception_type in {1, 5}:
            return self._record_action(
                workflow="exception-triage",
                action="auto_resolve",
                priority="P3-MEDIUM",
                requires_human_approval=False,
                summary=summary,
                rationale=(
                    "No-activity and expired-lot signals are low-risk "
                    "cleanup items in the tutorial guidance."
                ),
                evidence=evidence,
                payload={
                    "resolution": (
                        "Auto-resolved as low-risk operational housekeeping."
                    )
                },
            )

        return self._record_action(
            workflow="exception-triage",
            action="add_to_worklist",
            priority="P3-MEDIUM",
            requires_human_approval=True,
            summary=summary,
            rationale=(
                "Signal did not cross an auto-action threshold, so it "
                "should be queued for buyer review."
            ),
            evidence=evidence,
            payload={"worklist_bucket": self._exception_label(exception_type)},
        )

    def _get_supplier_on_time_rate(self, vendor_id: int) -> dict[str, Any]:
        """Calculate supplier on-time delivery performance."""
        try:
            sql = """
                SELECT
                    ph.VENDOR_ID,
                    COUNT(*) as total_deliveries,
                    SUM(CASE WHEN pll.ACTUAL_RECEIPT_DATE <= pll.NEED_BY_DATE
                             THEN 1 ELSE 0 END) as on_time_deliveries,
                    ROUND(100.0 * SUM(CASE WHEN pll.ACTUAL_RECEIPT_DATE <= pll.NEED_BY_DATE
                             THEN 1 ELSE 0 END) / COUNT(*), 1) as on_time_rate_pct,
                    SUM(CASE WHEN pll.ACTUAL_RECEIPT_DATE > pll.NEED_BY_DATE
                             THEN TRUNC(pll.ACTUAL_RECEIPT_DATE - pll.NEED_BY_DATE)
                             ELSE 0 END) as total_late_days
                FROM PO.PO_HEADERS_ALL ph
                JOIN PO.PO_LINE_LOCATIONS_ALL pll ON pll.PO_HEADER_ID = ph.PO_HEADER_ID
                WHERE ph.VENDOR_ID = :vendor_id
                AND ph.AUTHORIZATION_STATUS IN ('APPROVED', 'CLOSED')
                AND pll.ACTUAL_RECEIPT_DATE >= TRUNC(SYSDATE) - 365
                GROUP BY ph.VENDOR_ID
            """
            result = self.gateway.execute_query(
                sql,
                binds={"vendor_id": vendor_id},
                max_rows=1
            )
            return result[0] if result else {}
        except Exception:
            return {}

    def _get_alternate_suppliers(self, item_id: int, exclude_vendor_id: int | None = None) -> list[dict[str, Any]]:
        """Get alternate suppliers for an item, ranked by reliability."""
        try:
            sql = """
                SELECT mis.SUPPLIER_ID,
                       pv.VENDOR_NAME,
                       mis.LEAD_TIME,
                       mis.MIN_ORDER_QTY,
                       mis.SUPPLIER_ITEM_NUMBER,
                       ROW_NUMBER() OVER (ORDER BY mis.LEAD_TIME ASC, pv.VENDOR_NAME ASC) as rank
                FROM MSC.MSC_ITEM_SUPPLIERS mis
                JOIN PO.PO_VENDORS pv ON pv.VENDOR_ID = mis.SUPPLIER_ID
                WHERE mis.INVENTORY_ITEM_ID = :item_id
                AND mis.EFFECTIVE_DATE <= TRUNC(SYSDATE)
                AND (mis.DISABLE_DATE IS NULL OR mis.DISABLE_DATE > TRUNC(SYSDATE))
                AND (:exclude_vendor_id IS NULL OR mis.SUPPLIER_ID != :exclude_vendor_id)
                ORDER BY LEAD_TIME ASC
            """
            result = self.gateway.execute_query(
                sql,
                binds={"item_id": item_id,
                       "exclude_vendor_id": exclude_vendor_id},
                max_rows=5
            )
            return result if result else []
        except Exception:
            return []

    def run_late_supplier_detection(self, limit: int) -> dict[str, Any]:
        """Enhanced late supplier detection with chronic lateness scoring."""
        rows = self.gateway.get_late_supplier_candidates(
            days_ahead=3,
            limit=limit,
        )
        findings = []
        for row in rows:
            days_overdue = float(row.get("days_overdue") or 0)
            qty_outstanding = float(row.get("qty_outstanding") or 0)
            vendor_id = row.get("vendor_id")
            po_number = row.get("po_number")

            # Get supplier performance metrics
            supplier_perf = self._get_supplier_on_time_rate(
                vendor_id) if vendor_id else {}
            on_time_rate = float(supplier_perf.get("on_time_rate_pct") or 100)
            total_deliveries = int(supplier_perf.get("total_deliveries") or 0)

            # Use decision engine to evaluate supplier switch
            item_id = row.get("item_id")
            alternate_suppliers = []
            if item_id and (on_time_rate < 80 or total_deliveries > 3):
                alternate_suppliers = self._get_alternate_suppliers(
                    item_id, exclude_vendor_id=vendor_id)

            switch_decision = self.decision_engine.decide_supplier_switch(
                current_vendor={"vendor_name": row.get(
                    "vendor_name"), "vendor_id": vendor_id},
                current_performance=supplier_perf,
                alternates=alternate_suppliers,
            )

            is_chronic_late = switch_decision["decision"] == "switch"

            # Determine action based on decision engine output
            if is_chronic_late:
                action = "notify_buyer"
                priority = "P1-CRITICAL"
                summary = (
                    f"CHRONIC LATE SUPPLIER: PO {po_number} - "
                    f"Supplier {row.get('vendor_name')} has {on_time_rate:.0f}% on-time rate"
                )
                rationale = switch_decision["rationale"]
            elif days_overdue > 7 or qty_outstanding > 100:
                action = "notify_buyer"
                priority = "P1-CRITICAL"
                summary = f"PO {po_number} is overdue for supplier {row.get('vendor_name')}"
                rationale = "Severe overdue situation (7+ days or qty >100). Immediate buyer notification required."
            else:
                # Tutorial: First Offence → "Send reminder alert"
                action = "send_reminder_alert"
                priority = "P3-MEDIUM"
                summary = (
                    f"FIRST OFFENCE: PO {po_number} - "
                    f"Supplier {row.get('vendor_name')} approaching due date"
                )
                rationale = (
                    "First-time lateness detected. Tutorial prescribes sending a "
                    "reminder alert to the supplier before escalating."
                )

            if action == "send_reminder_alert":
                payload = {
                    "recommended_next_step": "Send reminder alert to supplier",
                    "alert_type": "first_offence",
                    "po_number": po_number,
                    "vendor_name": row.get("vendor_name"),
                    "days_overdue": days_overdue,
                    "qty_outstanding": qty_outstanding,
                    "message_template": (
                        f"Dear {row.get('vendor_name')}, "
                        f"PO {po_number} is approaching/past its due date with "
                        f"{qty_outstanding:.0f} units outstanding. "
                        "Please confirm delivery schedule immediately."
                    ),
                    "supplier_performance": {
                        "on_time_rate_pct": on_time_rate,
                        "total_deliveries": total_deliveries,
                        "is_chronic_late": False,
                    },
                    "decision_engine": switch_decision,
                }
            else:
                payload = {
                    "recommended_next_step": (
                        f"Switch to {switch_decision.get('recommended_vendor', {}).get('vendor_name', 'alternate')} (see list)"
                        if is_chronic_late and alternate_suppliers
                        else "Escalate supplier to procurement manager"
                    ),
                    "supplier_performance": {
                        "on_time_rate_pct": on_time_rate,
                        "total_deliveries": total_deliveries,
                        "is_chronic_late": is_chronic_late,
                        "total_late_days": supplier_perf.get("total_late_days", 0),
                    },
                    "decision_engine": switch_decision,
                    "alternate_suppliers": alternate_suppliers[:3],
                }

            findings.append(
                self._record_action(
                    workflow="late-supplier",
                    action=action,
                    priority=priority,
                    requires_human_approval=(action != "send_reminder_alert"),
                    summary=summary,
                    rationale=rationale,
                    evidence={**row, **supplier_perf},
                    payload=payload,
                    confidence=switch_decision.get("confidence"),
                )
            )
        return {
            "candidates_reviewed": len(rows),
            "chronic_late_suppliers_flagged": sum(1 for r in findings if "CHRONIC" in r.get("summary", "")),
            "sample_findings": findings[: min(10, len(findings))],
        }

    def run_safety_stock_alerts(
        self,
        plan_id: int,
        limit: int,
    ) -> dict[str, Any]:
        """Enhanced safety stock workflow: both shortage AND excess detection."""
        # Query for items BELOW safety stock (shortage)
        shortage_rows = self.gateway.execute_query(
            """
            SELECT *
            FROM (
                SELECT ss.INVENTORY_ITEM_ID,
                       ss.SAFETY_STOCK_QUANTITY AS target_qty,
                       NVL(s.current_supply, 0) AS current_supply,
                       ss.SAFETY_STOCK_QUANTITY - NVL(s.current_supply, 0) AS shortage,
                       'SHORTAGE' AS alert_type
                FROM MSC.MSC_SAFETY_STOCKS ss
                LEFT JOIN (
                    SELECT PLAN_ID,
                           INVENTORY_ITEM_ID,
                           SUM(NEW_ORDER_QUANTITY) AS current_supply
                    FROM MSC.MSC_SUPPLIES
                    WHERE ORDER_TYPE IN (1, 2, 3)
                    GROUP BY PLAN_ID, INVENTORY_ITEM_ID
                ) s ON s.PLAN_ID = ss.PLAN_ID AND s.INVENTORY_ITEM_ID = ss.INVENTORY_ITEM_ID
                WHERE ss.PLAN_ID = :plan_id
                AND NVL(s.current_supply, 0) < ss.SAFETY_STOCK_QUANTITY
                ORDER BY ss.SAFETY_STOCK_QUANTITY - NVL(s.current_supply, 0) DESC
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"plan_id": plan_id, "limit": limit},
            max_rows=limit,
        )

        # Query for items ABOVE safety stock (excess inventory)
        excess_rows = self.gateway.execute_query(
            """
            SELECT *
            FROM (
                SELECT ss.INVENTORY_ITEM_ID,
                       ss.SAFETY_STOCK_QUANTITY AS target_qty,
                       NVL(s.current_supply, 0) AS current_supply,
                       NVL(s.current_supply, 0) - ss.SAFETY_STOCK_QUANTITY AS excess,
                       'EXCESS' AS alert_type
                FROM MSC.MSC_SAFETY_STOCKS ss
                LEFT JOIN (
                    SELECT PLAN_ID,
                           INVENTORY_ITEM_ID,
                           SUM(NEW_ORDER_QUANTITY) AS current_supply
                    FROM MSC.MSC_SUPPLIES
                    WHERE ORDER_TYPE IN (1, 2, 3)
                    GROUP BY PLAN_ID, INVENTORY_ITEM_ID
                ) s ON s.PLAN_ID = ss.PLAN_ID AND s.INVENTORY_ITEM_ID = ss.INVENTORY_ITEM_ID
                WHERE ss.PLAN_ID = :plan_id
                AND NVL(s.current_supply, 0) > ss.SAFETY_STOCK_QUANTITY * 1.5
                ORDER BY NVL(s.current_supply, 0) - ss.SAFETY_STOCK_QUANTITY DESC
            )
            WHERE ROWNUM <= :limit
            """,
            binds={"plan_id": plan_id, "limit": limit},
            max_rows=limit,
        )

        findings = []

        # Process shortage alerts
        for row in shortage_rows:
            item_id = int(row.get("inventory_item_id") or 0)
            shortage_qty = float(row.get("shortage") or 0)
            suppliers = self.gateway.get_supplier_options(item_id)
            payload = self._create_draft_po_payload(
                item_id=item_id,
                quantity=shortage_qty,
                need_by_date=None,
                reason="Safety stock shortage detected - urgent replenishment required.",
                supplier_options=suppliers,
            )
            findings.append(
                self._action_or_recommendation(
                    workflow="safety-stock",
                    priority="P2-HIGH",
                    summary=(
                        f"Item {item_id} BELOW SAFETY STOCK by {shortage_qty:,.2f} units"
                    ),
                    rationale=(
                        "Item has fallen below safety stock threshold. "
                        "Urgent replenishment order recommended to prevent stockout."
                    ),
                    evidence=row,
                    payload=payload,
                    preferred_action="create_draft_po",
                )
            )

        # Process excess inventory alerts
        for row in excess_rows:
            item_id = int(row.get("inventory_item_id") or 0)
            excess_qty = float(row.get("excess") or 0)
            findings.append(
                self._record_action(
                    workflow="safety-stock",
                    action="flag_excess_inventory",
                    priority="P3-MEDIUM",
                    requires_human_approval=False,
                    summary=(
                        f"Item {item_id} has excess inventory of {excess_qty:,.2f} units above safety stock"
                    ),
                    rationale=(
                        "Item inventory is significantly above safety stock level. "
                        "Recommend review for: overstocking, slow-moving items, clearance sales, or donations."
                    ),
                    evidence=row,
                    payload={
                        "excess_quantity": excess_qty,
                        "target_quantity": row.get("target_qty"),
                        "current_quantity": row.get("current_supply"),
                        "recommended_actions": [
                            "Review demand forecast for accuracy",
                            "Consider clearance or promotional sales",
                            "Check for aged/obsolete inventory",
                            "Evaluate donation or disposal options"
                        ]
                    },
                )
            )

        return {
            "items_below_safety_stock": len(shortage_rows),
            "items_with_excess_inventory": len(excess_rows),
            "total_items_reviewed": len(shortage_rows) + len(excess_rows),
            "sample_findings": findings[: min(10, len(findings))],
        }

    def _get_contract_price(self, vendor_id: int, item_id: int) -> dict[str, Any]:
        """Look up contract/blanket agreement price for supplier-item combination."""
        try:
            sql = """
                SELECT
                    pa.AGREEMENT_NUM,
                    pa.AGREEMENT_TYPE,
                    pal.ITEM_ID,
                    pal.UNIT_PRICE AS contract_price,
                    pal.EFFECTIVE_DATE,
                    pal.EXPIRATION_DATE,
                    CASE WHEN pal.EXPIRATION_DATE < TRUNC(SYSDATE) THEN 'EXPIRED'
                         WHEN pal.EFFECTIVE_DATE > TRUNC(SYSDATE) THEN 'FUTURE'
                         ELSE 'ACTIVE' END AS contract_status
                FROM PO.PO_AGREEMENTS pa
                JOIN PO.PO_AGREEMENT_LINES pal ON pal.AGREEMENT_ID = pa.AGREEMENT_ID
                WHERE pa.VENDOR_ID = :vendor_id
                AND pal.ITEM_ID = :item_id
                AND pa.AUTHORIZATION_STATUS = 'APPROVED'
                ORDER BY pal.EXPIRATION_DATE DESC
            """
            result = self.gateway.execute_query(
                sql,
                binds={"vendor_id": vendor_id, "item_id": item_id},
                max_rows=1
            )
            return result[0] if result else {}
        except Exception:
            return {}

    def run_price_anomaly_detection(self, limit: int) -> dict[str, Any]:
        """Enhanced price anomaly detection with contract validation."""
        rows = self.gateway.get_price_anomalies(limit=limit)
        findings = []
        contracts_not_found = 0
        overpriced_items = 0

        for row in rows:
            current_price = float(row.get("unit_price") or 0)
            vendor_id = row.get("vendor_id")
            item_id = row.get("item_id")
            pct_deviation = float(row.get("pct_deviation") or 0)

            # Check if there's a contract price for this supplier-item
            contract_info = self._get_contract_price(vendor_id, item_id)
            contract_price = contract_info.get("contract_price")
            contract_status = contract_info.get("contract_status", "NONE")

            # Determine if price is overpriced vs contract
            is_over_contract = False
            renegotiation_savings = None
            if contract_price and contract_status == "ACTIVE":
                price_variance = (
                    (current_price - contract_price) / contract_price) * 100
                is_over_contract = price_variance > 5  # More than 5% above contract
                if is_over_contract:
                    renegotiation_savings = current_price - contract_price
                    overpriced_items += 1

            # Determine action based on severity and contract status
            if is_over_contract:
                action_priority = "P2-HIGH"
                action_summary = (
                    f"PO {row.get('po_number')} item {item_id}: "
                    f"Price ${current_price:.2f} exceeds active contract ${contract_price:.2f} by "
                    f"${renegotiation_savings:.2f}"
                )
                action_rationale = (
                    f"Current price deviates from active contract. "
                    f"Recommend renegotiation to enforce contract terms "
                    f"(potential savings: ${renegotiation_savings:.2f} per unit)"
                )
            elif contract_status == "EXPIRED" or contract_status == "NONE":
                action_priority = "P2-HIGH"
                action_summary = (
                    f"PO {row.get('po_number')} item {item_id}: "
                    f"Price deviation {pct_deviation}% and no active contract"
                )
                action_rationale = (
                    f"Price deviates by {pct_deviation}% from 12-month average. "
                    f"{'Active contract has expired - ' if contract_status == 'EXPIRED' else ''}"
                    f"Recommend establishing/renewing blanket agreement to lock pricing."
                )
                contracts_not_found += 1
            else:
                action_priority = "P2-HIGH"
                action_summary = (
                    f"PO {row.get('po_number')} item {item_id}: "
                    f"Price deviation {pct_deviation}%"
                )
                action_rationale = (
                    f"Price deviates {pct_deviation}% from historical average. "
                    f"Future contract available ({contract_status}). "
                    f"Recommend review for market conditions."
                )

            findings.append(
                self._record_action(
                    workflow="price-anomaly",
                    action="flag_for_renegotiation" if is_over_contract else "notify_buyer",
                    priority=action_priority,
                    requires_human_approval=True,
                    summary=action_summary,
                    rationale=action_rationale,
                    evidence={**row, **contract_info},
                    payload={
                        "current_price": current_price,
                        "contract_price": contract_price,
                        "contract_status": contract_status,
                        "price_deviation_pct": pct_deviation,
                        "renegotiation_potential_savings": renegotiation_savings,
                        "recommended_action": (
                            "Renegotiate PO to contract price" if is_over_contract
                            else "Establish or renew blanket agreement"
                        ),
                        "agreement_details": contract_info if contract_info else None
                    },
                )
            )

        return {
            "anomalies_found": len(rows),
            "items_over_contract_price": overpriced_items,
            "items_without_active_contract": contracts_not_found,
            "sample_findings": findings[: min(10, len(findings))],
        }

    def run_demand_to_po_tracing(
        self,
        plan_id: int,
        limit: int,
    ) -> dict[str, Any]:
        rows = self.gateway.get_demand_to_po_gaps(plan_id=plan_id, limit=limit)
        findings = []
        for row in rows:
            item_id = int(row.get("inventory_item_id") or 0)
            open_po = self.gateway.get_open_po_coverage(item_id, limit=3)
            if open_po:
                findings.append(
                    self._record_action(
                        workflow="demand-to-po",
                        action="monitor_delivery",
                        priority="P2-HIGH",
                        requires_human_approval=True,
                        summary=(
                            f"Planned order for item {item_id} already has "
                            "open PO coverage."
                        ),
                        rationale=(
                            "The tutorial says when a real PO already exists, "
                            "the agent should monitor delivery rather than "
                            "create another order."
                        ),
                        evidence={
                            "planned_order": row,
                            "open_po": open_po[:1],
                        },
                        payload={
                            "recommended_next_step": (
                                "Track delivery against demand."
                            )
                        },
                    )
                )
                continue

            suppliers = self.gateway.get_supplier_options(item_id)
            payload = self._create_draft_po_payload(
                item_id=item_id,
                quantity=float(row.get("planned_qty") or 0),
                need_by_date=row.get("need_by"),
                reason=(
                    "Demand-to-PO tracing found planned supply with no "
                    "real PO coverage."
                ),
                supplier_options=suppliers,
            )
            findings.append(
                self._action_or_recommendation(
                    workflow="demand-to-po",
                    priority="P1-CRITICAL",
                    summary=(
                        f"Planned purchase order for item {item_id} has no "
                        "real PO yet."
                    ),
                    rationale=(
                        "The tutorial says planned purchase orders without "
                        "corresponding real POs should become draft POs for "
                        "buyer approval."
                    ),
                    evidence=row,
                    payload=payload,
                    preferred_action="create_draft_po",
                )
            )
        return {
            "pegging_records_reviewed": len(rows),
            "sample_findings": findings[: min(10, len(findings))],
        }

    def run_spend_analytics(self, limit: int) -> dict[str, Any]:
        """Enhanced spend analytics: concentration risk, maverick spend, single-source, consolidation, time-period trends."""
        spend_rows = self.gateway.get_spend_summary(limit=limit)
        maverick_rows = self.gateway.get_maverick_spend(limit=limit)
        single_source_rows = self.gateway.get_single_source_items(limit=limit)
        consolidation_rows = self.gateway.get_consolidation_opportunities(
            limit=limit)
        time_period_rows = self.gateway.get_spend_by_time_period(limit=limit)
        findings = []

        # --- Concentration Risk Analysis ---
        total_spend = sum(float(r.get("total_spend") or 0) for r in spend_rows)
        top3_spend = sum(float(r.get("total_spend") or 0)
                         for r in spend_rows[:3])
        concentration_pct = (top3_spend / total_spend *
                             100) if total_spend > 0 else 0
        for row in spend_rows[: min(5, len(spend_rows))]:
            vendor_spend = float(row.get("total_spend") or 0)
            vendor_share = (vendor_spend / total_spend *
                            100) if total_spend > 0 else 0
            distinct_items = int(row.get("distinct_items") or 0)
            priority = "P2-HIGH" if vendor_share > 30 else "P3-MEDIUM"
            findings.append(
                self._record_action(
                    workflow="spend-analytics",
                    action="report_insight",
                    priority=priority,
                    requires_human_approval=False,
                    summary=(
                        f"Supplier {row.get('vendor_name')} accounts for "
                        f"{vendor_share:.1f}% of total spend across {distinct_items} items"
                    ),
                    rationale=(
                        "High supplier concentration increases supply chain risk. "
                        "Diversification and contract coverage should be reviewed."
                    ),
                    evidence=row,
                    payload={
                        "vendor_spend_share_pct": round(vendor_share, 2),
                        "total_portfolio_spend": round(total_spend, 2),
                        "top3_concentration_pct": round(concentration_pct, 2),
                        "recommended_next_step": (
                            "Establish blanket agreement to lock in pricing."
                            if vendor_share > 20
                            else "Monitor spend trend quarterly."
                        ),
                    },
                )
            )

        # --- Maverick Spend Detection (off-contract purchases) ---
        for row in maverick_rows[: min(3, len(maverick_rows))]:
            line_spend = float(row.get("line_spend") or 0)
            priority = "P2-HIGH" if line_spend > 10000 else "P3-MEDIUM"
            findings.append(
                self._record_action(
                    workflow="spend-analytics",
                    action="flag_maverick_spend",
                    priority=priority,
                    requires_human_approval=True,
                    summary=(
                        f"Maverick spend: PO {row.get('po_number')} from "
                        f"{row.get('vendor_name')} (${line_spend:,.2f}) has no active blanket agreement"
                    ),
                    rationale=(
                        "Purchases made outside of active blanket/contract agreements "
                        "bypass negotiated pricing and compliance controls. "
                        "These represent unauthorized off-contract spend."
                    ),
                    evidence=row,
                    payload={
                        "po_number": row.get("po_number"),
                        "vendor_name": row.get("vendor_name"),
                        "line_spend": line_spend,
                        "recommended_actions": [
                            "Establish blanket agreement with this supplier",
                            "Route future orders through procurement approval",
                            "Investigate if this spend could be consolidated",
                        ],
                    },
                )
            )

        # --- Single Source Risk Items ---
        for row in single_source_rows[: min(3, len(single_source_rows))]:
            item_spend = float(row.get("total_spend") or 0)
            priority = "P2-HIGH" if item_spend > 50000 else "P3-MEDIUM"
            findings.append(
                self._record_action(
                    workflow="spend-analytics",
                    action="flag_single_source_risk",
                    priority=priority,
                    requires_human_approval=False,
                    summary=(
                        f"Item {row.get('inventory_item_id')} sourced from only one supplier "
                        f"({row.get('sole_supplier_name')}) - supply chain risk"
                    ),
                    rationale=(
                        "Single-source items are a supply chain vulnerability. "
                        "Disruption from this supplier would halt production. "
                        "Dual-sourcing or approved alternate supplier registration is recommended."
                    ),
                    evidence=row,
                    payload={
                        "item_id": row.get("inventory_item_id"),
                        "sole_supplier": row.get("sole_supplier_name"),
                        "annual_spend": item_spend,
                        "recommended_actions": [
                            "Qualify an alternate supplier for this item",
                            "Add to approved supplier list (ASL) in Oracle",
                            "Build safety stock to buffer single-source risk",
                        ],
                    },
                )
            )

        # --- Consolidation Opportunities ---
        for row in consolidation_rows[: min(3, len(consolidation_rows))]:
            price_spread = float(row.get("price_spread") or 0)
            min_price = float(row.get("min_price") or 0)
            max_price = float(row.get("max_price") or 0)
            total_spend = float(row.get("total_spend") or 0)
            spread_pct = ((max_price - min_price) / min_price *
                          100) if min_price > 0 else 0
            findings.append(
                self._record_action(
                    workflow="spend-analytics",
                    action="consolidate_spend",
                    priority="P3-MEDIUM",
                    requires_human_approval=False,
                    summary=(
                        f"Item {row.get('inventory_item_id')} bought from "
                        f"{row.get('supplier_count')} suppliers with {spread_pct:.1f}% price spread "
                        f"(${min_price:.2f}-${max_price:.2f})"
                    ),
                    rationale=(
                        "Multiple suppliers for the same item with significant price spread "
                        "represents a consolidation opportunity. Routing volume to lowest-cost "
                        "supplier could reduce spend."
                    ),
                    evidence=row,
                    payload={
                        "item_id": row.get("inventory_item_id"),
                        "supplier_count": row.get("supplier_count"),
                        "min_price": min_price,
                        "max_price": max_price,
                        "price_spread_pct": round(spread_pct, 2),
                        "total_spend": total_spend,
                        "recommended_actions": [
                            f"Consolidate volume to lowest-price supplier (save up to {spread_pct:.0f}%)",
                            "Negotiate volume discount with preferred supplier",
                            "Establish blanket agreement at negotiated price",
                        ],
                    },
                )
            )

        return {
            "suppliers_ranked": len(spend_rows),
            "total_spend": round(sum(float(r.get("total_spend") or 0) for r in spend_rows), 2),
            "top3_concentration_pct": round(concentration_pct, 2),
            "maverick_spend_count": len(maverick_rows),
            "single_source_items": len(single_source_rows),
            "consolidation_opportunities": len(consolidation_rows),
            "quarterly_periods_analysed": len(time_period_rows),
            "top_suppliers": spend_rows[: min(10, len(spend_rows))],
            "time_period_trends": time_period_rows,
            "sample_findings": findings,
        }


def _tool_schemas() -> list[dict[str, Any]]:
    return [
        {
            "name": "get_exception_summary",
            "description": "Get total exception count and top plans.",
            "input_schema": {
                "type": "object",
                "properties": {"limit": {"type": "integer"}},
            },
        },
        {
            "name": "get_exception_types",
            "description": "Get exception type breakdown for a plan.",
            "input_schema": {
                "type": "object",
                "properties": {"plan_id": {"type": "integer"}},
                "required": ["plan_id"],
            },
        },
        {
            "name": "get_exception_details",
            "description": "Get exception rows for a plan and exception type.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "plan_id": {"type": "integer"},
                    "exception_type": {"type": "integer"},
                    "limit": {"type": "integer"},
                },
                "required": ["plan_id", "exception_type"],
            },
        },
        {
            "name": "get_item_context",
            "description": "Get item master context for an inventory item.",
            "input_schema": {
                "type": "object",
                "properties": {"item_id": {"type": "integer"}},
                "required": ["item_id"],
            },
        },
        {
            "name": "get_supplier_options",
            "description": "Get supplier options for an item.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "item_id": {"type": "integer"},
                    "limit": {"type": "integer"},
                },
                "required": ["item_id"],
            },
        },
        {
            "name": "get_open_po_coverage",
            "description": "Get open PO coverage for an item.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "item_id": {"type": "integer"},
                    "limit": {"type": "integer"},
                },
                "required": ["item_id"],
            },
        },
        {
            "name": "get_safety_stock_context",
            "description": "Get safety stock context for an item and plan.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "item_id": {"type": "integer"},
                    "plan_id": {"type": "integer"},
                },
                "required": ["item_id", "plan_id"],
            },
        },
        {
            "name": "get_pegging_context",
            "description": "Get pegging context for an item and plan.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "item_id": {"type": "integer"},
                    "plan_id": {"type": "integer"},
                    "limit": {"type": "integer"},
                },
                "required": ["item_id", "plan_id"],
            },
        },
        {
            "name": "get_late_supplier_candidates",
            "description": "Get candidate late supplier PO lines.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "days_ahead": {"type": "integer"},
                    "limit": {"type": "integer"},
                },
            },
        },
        {
            "name": "get_price_anomalies",
            "description": (
                "Get PO price anomalies using historical baselines."
            ),
            "input_schema": {
                "type": "object",
                "properties": {"limit": {"type": "integer"}},
            },
        },
        {
            "name": "get_demand_to_po_gaps",
            "description": (
                "Get planned purchase orders without confirmed PO tracing."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "plan_id": {"type": "integer"},
                    "limit": {"type": "integer"},
                },
                "required": ["plan_id"],
            },
        },
        {
            "name": "get_spend_summary",
            "description": "Get supplier spend summary.",
            "input_schema": {
                "type": "object",
                "properties": {"limit": {"type": "integer"}},
            },
        },
        {
            "name": "create_draft_po",
            "description": "Create a draft PO payload for buyer review.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "workflow": {"type": "string"},
                    "item_id": {"type": "integer"},
                    "quantity": {"type": "number"},
                    "need_by_date": {"type": "string"},
                    "reason": {"type": "string"},
                    "priority": {"type": "string"},
                    "evidence": {"type": "object"},
                    "confidence": {"type": "number"},
                },
                "required": [
                    "workflow",
                    "item_id",
                    "quantity",
                    "reason",
                    "priority",
                    "evidence",
                ],
            },
        },
        {
            "name": "notify_buyer",
            "description": "Notify buyer for review or escalation.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "workflow": {"type": "string"},
                    "priority": {"type": "string"},
                    "summary": {"type": "string"},
                    "rationale": {"type": "string"},
                    "evidence": {"type": "object"},
                    "revenue_at_risk": {"type": "number"},
                    "confidence": {"type": "number"},
                    "assigned_to": {"type": "string"},
                },
                "required": [
                    "workflow",
                    "priority",
                    "summary",
                    "rationale",
                    "evidence",
                ],
            },
        },
        {
            "name": "add_to_worklist",
            "description": "Add an issue to the buyer worklist.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "workflow": {"type": "string"},
                    "priority": {"type": "string"},
                    "summary": {"type": "string"},
                    "rationale": {"type": "string"},
                    "evidence": {"type": "object"},
                    "bucket": {"type": "string"},
                    "confidence": {"type": "number"},
                },
                "required": [
                    "workflow",
                    "priority",
                    "summary",
                    "rationale",
                    "evidence",
                    "bucket",
                ],
            },
        },
        {
            "name": "auto_resolve",
            "description": "Auto-resolve a low-risk issue.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "workflow": {"type": "string"},
                    "priority": {"type": "string"},
                    "summary": {"type": "string"},
                    "rationale": {"type": "string"},
                    "evidence": {"type": "object"},
                    "resolution": {"type": "string"},
                    "confidence": {"type": "number"},
                },
                "required": [
                    "workflow",
                    "priority",
                    "summary",
                    "rationale",
                    "evidence",
                    "resolution",
                ],
            },
        },
        {
            "name": "report_insight",
            "description": (
                "Record a reporting insight when no action is required."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "workflow": {"type": "string"},
                    "summary": {"type": "string"},
                    "rationale": {"type": "string"},
                    "evidence": {"type": "object"},
                    "next_step": {"type": "string"},
                    "confidence": {"type": "number"},
                },
                "required": [
                    "workflow",
                    "summary",
                    "rationale",
                    "evidence",
                    "next_step",
                ],
            },
        },
    ]


def _build_tool_dispatch(
    gateway: OracleReadOnlyGateway,
    agent: TutorialProcurementAgent,
) -> dict[str, Any]:
    return {
        "get_exception_summary": lambda a: gateway.get_exception_summary(
            a.get("limit", 10)
        ),
        "get_exception_types": lambda a: gateway.get_exception_types(
            a["plan_id"]
        ),
        "get_exception_details": lambda a: gateway.get_exception_details(
            a["plan_id"],
            a["exception_type"],
            a.get("limit", 10),
        ),
        "get_item_context": lambda a: gateway.get_item_context(a["item_id"]),
        "get_supplier_options": lambda a: gateway.get_supplier_options(
            a["item_id"],
            a.get("limit", 5),
        ),
        "get_open_po_coverage": lambda a: gateway.get_open_po_coverage(
            a["item_id"],
            a.get("limit", 10),
        ),
        "get_safety_stock_context": lambda a: gateway.get_safety_stock_context(
            a["item_id"],
            a["plan_id"],
        ),
        "get_pegging_context": lambda a: gateway.get_pegging_context(
            a["item_id"],
            a["plan_id"],
            a.get("limit", 5),
        ),
        "get_late_supplier_candidates": (
            lambda a: gateway.get_late_supplier_candidates(
                a.get("days_ahead", 3),
                a.get("limit", 10),
            )
        ),
        "get_price_anomalies": lambda a: gateway.get_price_anomalies(
            a.get("limit", 10)
        ),
        "get_demand_to_po_gaps": lambda a: gateway.get_demand_to_po_gaps(
            a["plan_id"],
            a.get("limit", 10),
        ),
        "get_spend_summary": lambda a: gateway.get_spend_summary(
            a.get("limit", 10)
        ),
        "create_draft_po": lambda a: agent.create_draft_po_action(
            workflow=a["workflow"],
            item_id=a["item_id"],
            quantity=a["quantity"],
            need_by_date=a.get("need_by_date"),
            reason=a["reason"],
            priority=a["priority"],
            evidence=a["evidence"],
            confidence=a.get("confidence"),
        ),
        "notify_buyer": lambda a: agent.notify_buyer_action(
            workflow=a["workflow"],
            priority=a["priority"],
            summary=a["summary"],
            rationale=a["rationale"],
            evidence=a["evidence"],
            revenue_at_risk=a.get("revenue_at_risk"),
            confidence=a.get("confidence"),
            assigned_to=a.get("assigned_to", "Senior Buyer"),
        ),
        "add_to_worklist": lambda a: agent.add_to_worklist_action(
            workflow=a["workflow"],
            priority=a["priority"],
            summary=a["summary"],
            rationale=a["rationale"],
            evidence=a["evidence"],
            bucket=a["bucket"],
            confidence=a.get("confidence"),
        ),
        "auto_resolve": lambda a: agent.auto_resolve_action(
            workflow=a["workflow"],
            priority=a["priority"],
            summary=a["summary"],
            rationale=a["rationale"],
            evidence=a["evidence"],
            resolution=a["resolution"],
            confidence=a.get("confidence"),
        ),
        "report_insight": lambda a: agent.report_insight_action(
            workflow=a["workflow"],
            summary=a["summary"],
            rationale=a["rationale"],
            evidence=a["evidence"],
            next_step=a["next_step"],
            confidence=a.get("confidence"),
        ),
    }


def _build_system_prompt(
    workflow: str,
    autonomy_level: int,
    plan_id: int | None,
    limit: int,
) -> str:
    return f"""
You are an AI Procurement Agent implementing the tutorial in
TUTORIAL_Agentic_Procurement_AI.md.

Architecture to follow:
- Read Oracle EBS data through tools only.
- Use planning and procurement context before acting.
- Respect autonomy level {autonomy_level}.
- Never write live Oracle transactions.
- Draft POs are payloads for buyer approval only.
- Log every action with rationale, evidence, and confidence.

Autonomy policy:
- Level 0: Inform only.
- Level 1: Recommend actions.
- Level 2: Create draft PO payloads and other non-destructive actions.

Workflow requested: {workflow}
Plan override: {plan_id}
Record limit guidance: {limit}

When you act:
- Use create_draft_po for shortage and past due procurement actions.
- Use notify_buyer for complex or demand-facing risk.
- Use add_to_worklist for medium-priority cases.
- Use auto_resolve for low-risk housekeeping.
- Use report_insight for analytics-only findings.

Priorities:
- P1-CRITICAL
- P2-HIGH
- P3-MEDIUM

You must finish with a concise final summary after taking actions.
""".strip()


def _build_user_prompt(
    workflow: str,
    autonomy_level: int,
    plan_id: int | None,
    limit: int,
) -> str:
    return (
        f"Run the {workflow} workflow from the tutorial. "
        f"Autonomy level is {autonomy_level}. "
        f"Use up to {limit} records per exploration step. "
        f"If a plan is needed, use plan_id {plan_id} if provided, otherwise "
        "discover the most relevant plan from exception volume."
    )


def _run_claude_agent(
    gateway: OracleReadOnlyGateway,
    agent: TutorialProcurementAgent,
    workflow: str,
    plan_id: int | None,
    limit: int,
) -> dict[str, Any]:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError(
            "Set ANTHROPIC_API_KEY in .env or your terminal environment "
            "to use --engine claude."
        )

    model = os.getenv("ANTHROPIC_MODEL", "claude-opus-4-6")
    client = anthropic.Anthropic(api_key=api_key)
    tools = _tool_schemas()
    dispatch = _build_tool_dispatch(gateway, agent)

    messages: list[dict[str, Any]] = [
        {"role": "user", "content": _build_user_prompt(
            workflow, agent.autonomy_level, plan_id, limit)}
    ]
    transcript: list[dict[str, Any]] = []

    for turn in range(1, 21):
        response = client.messages.create(
            model=model,
            max_tokens=4096,
            system=_build_system_prompt(
                workflow, agent.autonomy_level, plan_id, limit),
            tools=tools,
            messages=messages,
        )

        messages.append({"role": "assistant", "content": response.content})
        transcript.append(
            {
                "turn": turn,
                "stop_reason": response.stop_reason,
                "content": [block.model_dump() for block in response.content],
            }
        )

        tool_blocks = [
            block for block in response.content if block.type == "tool_use"]
        if not tool_blocks:
            text_blocks = [
                block.text
                for block in response.content
                if block.type == "text"
            ]
            return {
                "engine": "claude",
                "model": model,
                "transcript": transcript,
                "final_text": "\n".join(text_blocks).strip(),
                "action_summary": dict(
                    Counter(action.action for action in agent.actions)
                ),
                "actions": [asdict(action) for action in agent.actions],
            }

        tool_results = []
        for block in tool_blocks:
            fn = dispatch[block.name]
            result = fn(block.input)
            tool_results.append(
                {
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": json.dumps(result, default=_json_default),
                }
            )
        messages.append({"role": "user", "content": tool_results})

    return {
        "engine": "claude",
        "model": model,
        "transcript": transcript,
        "final_text": "Claude agent reached max turns.",
        "action_summary": dict(
            Counter(action.action for action in agent.actions)
        ),
        "actions": [asdict(action) for action in agent.actions],
    }


def _save_excel(report: dict[str, Any]) -> Path:  # noqa: C901
    """Generate a professional, fully-formatted Excel workbook from the agent run report.

    Sheets:
      1. Dashboard   – Run summary + action count KPIs + workflow metrics table
      2. All Actions – One row per action, color-coded by priority, structured columns
      3. Draft POs   – Structured PO data (header/line/shipment/distribution columns)
      4. Spend Analytics – Top suppliers + maverick/single-source flags
      5. Workflow Summary – Key metric table per workflow (no JSON dumps)
    """
    from openpyxl.styles import Alignment, Border, Side, numbers as xl_numbers
    from openpyxl.utils import get_column_letter

    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = report.get(
        "run_id") or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"tutorial_agent_run_{timestamp}.xlsx"

    result = report.get("result", {})
    actions = result.get("actions", [])
    workflows = result.get("workflows", {})
    conn_info = report.get("connection", {})

    # ── Colour palette ──────────────────────────────────────────────────────
    C_HDR_DARK = "1F4E78"   # deep navy  – sheet headers
    C_HDR_MID = "2E75B6"   # mid-blue   – sub-headers
    C_P1 = "FF0000"   # red fill   – P1-CRITICAL
    C_P1_TXT = "FFFFFF"
    C_P2 = "ED7D31"   # orange     – P2-HIGH
    C_P2_TXT = "FFFFFF"
    C_P3 = "FFD966"   # amber      – P3-MEDIUM
    C_P3_TXT = "000000"
    C_P4 = "A9D18E"   # green      – P4-LOW / insight
    C_P4_TXT = "000000"
    C_ROW_ALT = "DEEAF1"   # light-blue – alternating rows
    C_LABEL_BG = "D6E4F0"   # pale-blue  – label cells on dashboard
    C_WHITE = "FFFFFF"
    C_BORDER = "BFBFBF"

    PRIORITY_FILLS = {
        "P1-CRITICAL": (C_P1, C_P1_TXT),
        "P2-HIGH":     (C_P2, C_P2_TXT),
        "P3-MEDIUM":   (C_P3, C_P3_TXT),
        "P4-LOW":      (C_P4, C_P4_TXT),
    }

    def hfill(color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    def hfont(color: str = "FFFFFF", bold: bool = True, size: int = 10) -> Font:
        return Font(color=color, bold=bold, size=size, name="Arial")

    def border(color: str = C_BORDER) -> Border:
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def center(wrap: bool = False) -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left(wrap: bool = False) -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    def _write_header_row(ws, row: int, cols: list[str], fill_color: str = C_HDR_DARK) -> None:
        for c, label in enumerate(cols, 1):
            cell = ws.cell(row, c, label)
            cell.fill = hfill(fill_color)
            cell.font = hfont("FFFFFF", bold=True, size=10)
            cell.border = border()
            cell.alignment = center(wrap=True)

    def _data_cell(ws, row: int, col: int, value: Any,
                   fill: str | None = None, txt_color: str = "000000",
                   bold: bool = False, wrap: bool = False, align: str = "left") -> None:
        cell = ws.cell(row, col, _safe(value))
        if fill:
            cell.fill = hfill(fill)
        cell.font = Font(color=txt_color, bold=bold, size=10, name="Arial")
        cell.border = border()
        cell.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=wrap)

    def _safe(v: Any) -> Any:
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return json.dumps(v, default=_json_default)
        if isinstance(v, datetime):
            return v.isoformat()
        return v

    def _set_col_widths(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _priority_colors(priority: str) -> tuple[str, str]:
        return PRIORITY_FILLS.get(priority, (C_P4, C_P4_TXT))

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 – DASHBOARD
    # ════════════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False

    # Title banner
    ws_dash.merge_cells("A1:F1")
    title_cell = ws_dash["A1"]
    title_cell.value = "🤖  Agentic Procurement AI  –  Run Report"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = hfill(C_HDR_DARK)
    title_cell.alignment = center()
    ws_dash.row_dimensions[1].height = 36

    # Sub-title
    ws_dash.merge_cells("A2:F2")
    sub = ws_dash["A2"]
    sub.value = f"Generated: {report.get('generated_at', '')}   |   DB: {conn_info.get('sid', '')}   |   User: {conn_info.get('connected_user', '')}   |   Version: {conn_info.get('db_version', '')}"
    sub.font = Font(name="Arial", size=9, color="FFFFFF")
    sub.fill = hfill(C_HDR_MID)
    sub.alignment = center()
    ws_dash.row_dimensions[2].height = 18

    # ── Run details block ──
    run_labels = [
        ("Tutorial Source",   report.get("tutorial_source", "")),
        ("Engine",            result.get("engine", "rules")),
        ("Autonomy Level",    result.get("autonomy_level", "")),
        ("Selected Plan ID",  result.get("selected_plan_id", "")),
        ("Selected Plan Name", result.get("selected_plan_name", "")),
    ]
    for i, (label, value) in enumerate(run_labels, start=4):
        lc = ws_dash.cell(i, 1, label)
        lc.fill = hfill(C_LABEL_BG)
        lc.font = hfont("000000", bold=True, size=10)
        lc.border = border()
        lc.alignment = left()
        vc = ws_dash.cell(i, 2, _safe(value))
        vc.font = Font(name="Arial", size=10)
        vc.border = border()
        vc.alignment = left()

    # ── KPI boxes – action counts ──
    ws_dash.merge_cells("D4:F4")
    kpi_hdr = ws_dash["D4"]
    kpi_hdr.value = "Action Count KPIs"
    kpi_hdr.fill = hfill(C_HDR_MID)
    kpi_hdr.font = hfont("FFFFFF", bold=True)
    kpi_hdr.alignment = center()
    kpi_hdr.border = border()

    action_summary = result.get("action_summary", {})
    kpi_row = 5
    for action_name, count in sorted(action_summary.items()):
        ws_dash.cell(kpi_row, 4, action_name).border = border()
        ws_dash.cell(kpi_row, 4).font = Font(name="Arial", size=10)
        ws_dash.cell(kpi_row, 4).alignment = left()
        cnt_cell = ws_dash.cell(kpi_row, 5, count)
        cnt_cell.font = Font(name="Arial", size=10, bold=True)
        cnt_cell.alignment = center()
        cnt_cell.border = border()
        ws_dash.merge_cells(f"D{kpi_row}:E{kpi_row}")
        ws_dash.cell(kpi_row, 6, count).border = border()
        ws_dash.cell(kpi_row, 6).font = Font(name="Arial", size=11, bold=True)
        ws_dash.cell(kpi_row, 6).alignment = center()
        kpi_row += 1

    # ── Workflow metrics table ──
    wf_start = max(kpi_row, len(run_labels) + 5) + 1
    ws_dash.merge_cells(f"A{wf_start}:F{wf_start}")
    wf_hdr = ws_dash.cell(wf_start, 1, "Workflow Execution Summary")
    wf_hdr.fill = hfill(C_HDR_DARK)
    wf_hdr.font = hfont("FFFFFF", bold=True)
    wf_hdr.alignment = center()
    wf_hdr.border = border()

    wf_cols = ["Workflow", "Records / Items",
               "Actions Created", "Key Metric", "Status"]
    _write_header_row(ws_dash, wf_start + 1, wf_cols, C_HDR_MID)

    WF_METRICS = {
        "exception-triage":  ("decisions_made",          "exception_types"),
        "late-supplier":     ("candidates_reviewed",     "chronic_late_suppliers_flagged"),
        "safety-stock":      ("total_items_reviewed",    "items_below_safety_stock"),
        "price-anomaly":     ("anomalies_found",         "items_over_contract_price"),
        "demand-to-po":      ("pegging_records_reviewed", None),
        "spend-analytics":   ("suppliers_ranked",        "maverick_spend_count"),
    }

    for ri, (wf_name, wf_data) in enumerate(workflows.items(), start=wf_start + 2):
        rec_key, metric_key = WF_METRICS.get(wf_name, ("?", None))
        records = wf_data.get(rec_key, "—")
        actions_count = sum(1 for a in actions if a.get("workflow") == wf_name)
        key_metric = wf_data.get(metric_key, "—") if metric_key else "—"
        status = "✅ Done" if records not in ("—", 0, None) else "⚠ No data"
        row_fill = C_ROW_ALT if ri % 2 == 0 else C_WHITE
        for ci, val in enumerate([wf_name, records, actions_count, key_metric, status], 1):
            _data_cell(ws_dash, ri, ci, val, fill=row_fill,
                       align="center" if ci > 1 else "left")

    _set_col_widths(ws_dash, [24, 18, 16, 12, 14, 12])

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 – ALL ACTIONS (color-coded by priority)
    # ════════════════════════════════════════════════════════════════════════
    ws_act = wb.create_sheet("All Actions")
    ws_act.sheet_view.showGridLines = False

    # Freeze top row
    ws_act.freeze_panes = "A2"

    act_cols = [
        "#", "Workflow", "Action", "Priority", "Requires Approval",
        "Summary", "Rationale", "Confidence", "Assigned To",
    ]
    _write_header_row(ws_act, 1, act_cols)

    for ri, action in enumerate(actions, start=2):
        priority = action.get("priority", "P4-LOW")
        p_fill, p_txt = _priority_colors(priority)
        row_fill = C_ROW_ALT if ri % 2 == 0 else C_WHITE

        _data_cell(ws_act, ri, 1, ri - 1, fill=row_fill, align="center")
        _data_cell(ws_act, ri, 2, action.get("workflow", ""), fill=row_fill)
        _data_cell(ws_act, ri, 3, action.get("action", ""), fill=row_fill)
        # Priority cell – colored by severity
        _data_cell(ws_act, ri, 4, priority, fill=p_fill,
                   txt_color=p_txt, bold=True, align="center")
        approval = "YES" if action.get("requires_human_approval") else "No"
        _data_cell(ws_act, ri, 5, approval, fill=row_fill, align="center")
        _data_cell(ws_act, ri, 6, action.get(
            "summary", ""), fill=row_fill, wrap=True)
        _data_cell(ws_act, ri, 7, action.get(
            "rationale", ""), fill=row_fill, wrap=True)
        conf = action.get("confidence")
        _data_cell(
            ws_act, ri, 8, f"{conf:.0%}" if conf else "—", fill=row_fill, align="center")
        _data_cell(ws_act, ri, 9, action.get(
            "assigned_to", "—"), fill=row_fill)
        ws_act.row_dimensions[ri].height = 42

    _set_col_widths(ws_act, [5, 18, 22, 14, 16, 55, 55, 12, 18])

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3 – DRAFT POs (structured columns, not JSON)
    # ════════════════════════════════════════════════════════════════════════
    ws_po = wb.create_sheet("Draft POs")
    ws_po.sheet_view.showGridLines = False
    ws_po.freeze_panes = "A2"

    po_actions = [a for a in actions if a.get("action") in (
        "recommend_draft_po", "create_draft_po")]

    po_cols = [
        "PO Number", "Status", "Workflow", "Priority",
        # Header
        "Vendor ID", "Vendor Name", "Org ID", "Currency",
        "Payment Terms ID", "Ship-To Location ID", "FOB Code",
        # Line
        "Item ID", "Item Number", "Item Description", "UOM", "Quantity", "Unit Price", "Line Amount",
        # Distribution
        "Charge Account", "Destination Type", "Amount Ordered",
        # Validation
        "Can Create PO", "Supplier Status", "Item Status", "Warnings",
    ]
    _write_header_row(ws_po, 1, po_cols)

    for ri, action in enumerate(po_actions, start=2):
        payload = action.get("payload", {})
        hdr = payload.get("header", {})
        line = payload.get("line", {})
        ship = payload.get("shipment", {})
        dist = payload.get("distribution", {})
        vs = payload.get("validation_status", {})
        priority = action.get("priority", "P4-LOW")
        p_fill, p_txt = _priority_colors(priority)
        row_fill = C_ROW_ALT if ri % 2 == 0 else C_WHITE

        vals = [
            payload.get("po_number", ""),
            payload.get("status", ""),
            action.get("workflow", ""),
            priority,
            hdr.get("vendor_id", ""),
            hdr.get("vendor_name", ""),
            hdr.get("org_id", ""),
            hdr.get("currency_code", ""),
            hdr.get("payment_terms_id", ""),
            hdr.get("ship_to_location_id", ""),
            hdr.get("fob_lookup_code", ""),
            line.get("item_id", ""),
            line.get("item_number", ""),
            line.get("item_description", ""),
            line.get("unit_meas_lookup_code", ""),
            line.get("quantity", ""),
            line.get("unit_price", ""),
            line.get("line_amount", ""),
            dist.get("charge_account", ""),
            dist.get("destination_type_code", ""),
            dist.get("amount_ordered", ""),
            "YES" if vs.get("can_create_po") else "NO",
            vs.get("supplier_status", ""),
            vs.get("item_status", ""),
            "; ".join(vs.get("warnings", [])),
        ]
        for ci, val in enumerate(vals, 1):
            fill = p_fill if ci == 4 else row_fill
            txt = p_txt if ci == 4 else "000000"
            bld = ci == 4
            _data_cell(ws_po, ri, ci, val, fill=fill, txt_color=txt, bold=bld,
                       wrap=(ci in (14, 25)), align="center" if ci in (4, 6, 12, 22, 23, 24) else "left")
        ws_po.row_dimensions[ri].height = 36

    _set_col_widths(ws_po, [30, 22, 18, 14, 10, 22, 10, 10, 16, 18, 12,
                            10, 14, 28, 6, 10, 12, 14, 22, 16, 14, 12, 16, 16, 40])

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 4 – SPEND ANALYTICS
    # ════════════════════════════════════════════════════════════════════════
    ws_spend = wb.create_sheet("Spend Analytics")
    ws_spend.sheet_view.showGridLines = False
    sa_data = workflows.get("spend-analytics", {})

    # Section: Top Suppliers
    ws_spend.merge_cells("A1:F1")
    sec1 = ws_spend["A1"]
    sec1.value = "Top Suppliers by Spend"
    sec1.fill = hfill(C_HDR_DARK)
    sec1.font = hfont("FFFFFF", bold=True)
    sec1.alignment = center()
    sec1.border = border()

    sup_cols = ["Vendor ID", "Vendor Name", "PO Count",
                "Distinct Items", "Total Spend (USD)", "Currency"]
    _write_header_row(ws_spend, 2, sup_cols, C_HDR_MID)

    top_suppliers = sa_data.get("top_suppliers", [])
    for ri, sup in enumerate(top_suppliers, start=3):
        row_fill = C_ROW_ALT if ri % 2 == 1 else C_WHITE
        vals = [
            sup.get("vendor_id", ""),
            sup.get("vendor_name", ""),
            sup.get("po_count", ""),
            sup.get("distinct_items", ""),
            sup.get("total_spend", ""),
            sup.get("currency_code", ""),
        ]
        for ci, val in enumerate(vals, 1):
            _data_cell(ws_spend, ri, ci, val, fill=row_fill,
                       align="right" if ci == 5 else ("center" if ci in (1, 3, 4, 6) else "left"))

    # Section: Maverick Spend findings
    mav_start = len(top_suppliers) + 5
    ws_spend.merge_cells(f"A{mav_start}:F{mav_start}")
    sec2 = ws_spend.cell(
        mav_start, 1, "⚠  Maverick Spend / Single-Source / Consolidation Flags")
    sec2.fill = hfill(C_HDR_MID)
    sec2.font = hfont("FFFFFF", bold=True)
    sec2.alignment = center()
    sec2.border = border()

    flag_cols = ["Action Type", "Priority",
                 "Summary", "Rationale", "Recommended Actions"]
    _write_header_row(ws_spend, mav_start + 1, flag_cols, C_HDR_DARK)

    spend_actions = [a for a in actions if a.get(
        "workflow") == "spend-analytics"]
    for ri, action in enumerate(spend_actions, start=mav_start + 2):
        priority = action.get("priority", "P3-MEDIUM")
        p_fill, p_txt = _priority_colors(priority)
        row_fill = C_ROW_ALT if ri % 2 == 0 else C_WHITE
        rec_actions = action.get("payload", {}).get("recommended_actions", [])
        rec_str = "\n".join(f"• {r}" for r in rec_actions) if rec_actions else action.get(
            "payload", {}).get("recommended_next_step", "—")
        for ci, val in enumerate([
            action.get("action", ""),
            priority,
            action.get("summary", ""),
            action.get("rationale", ""),
            rec_str,
        ], 1):
            fill = p_fill if ci == 2 else row_fill
            txt = p_txt if ci == 2 else "000000"
            _data_cell(ws_spend, ri, ci, val, fill=fill, txt_color=txt,
                       bold=(ci == 2), wrap=(ci in (3, 4, 5)))
        ws_spend.row_dimensions[ri].height = 54

    _set_col_widths(ws_spend, [24, 14, 52, 50, 50])

    # Section: Quarterly Spend Trend
    time_period_rows = sa_data.get("time_period_trends", [])
    if time_period_rows:
        trend_start = len(top_suppliers) + len(spend_actions) + 8
        ws_spend.merge_cells(f"A{trend_start}:G{trend_start}")
        sec3 = ws_spend.cell(
            trend_start, 1, "Quarterly Spend Trend (Last 24 Months)")
        sec3.fill = hfill(C_HDR_MID)
        sec3.font = hfont("FFFFFF", bold=True)
        sec3.alignment = center()
        sec3.border = border()

        trend_cols = ["Quarter", "Year", "Business Unit", "PO Count",
                      "Vendors", "Items", "Total Spend (USD)", "Currency"]
        _write_header_row(ws_spend, trend_start + 1, trend_cols, C_HDR_DARK)

        for ri, tr in enumerate(time_period_rows, start=trend_start + 2):
            row_fill = C_ROW_ALT if ri % 2 == 1 else C_WHITE
            vals = [
                tr.get("time_period", ""),
                tr.get("year", ""),
                tr.get("business_unit_id", ""),
                tr.get("po_count", ""),
                tr.get("vendor_count", ""),
                tr.get("item_count", ""),
                tr.get("total_spend", ""),
                tr.get("currency_code", ""),
            ]
            for ci, val in enumerate(vals, 1):
                align = "right" if ci == 7 else (
                    "center" if ci in (1, 2, 4, 5, 6, 8) else "left")
                _data_cell(ws_spend, ri, ci, val, fill=row_fill, align=align)

        # extend column widths for 8 cols (trend section)
        for ci, w in enumerate([14, 8, 18, 10, 10, 10, 22, 10], 1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            if ws_spend.column_dimensions[col_letter].width < w:
                ws_spend.column_dimensions[col_letter].width = w

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 5 – WORKFLOW SUMMARY (structured metrics, no JSON)
    # ════════════════════════════════════════════════════════════════════════
    ws_wf = wb.create_sheet("Workflow Summary")
    ws_wf.sheet_view.showGridLines = False

    ws_wf.merge_cells("A1:C1")
    wf_title = ws_wf["A1"]
    wf_title.value = "Workflow Results – Key Metrics"
    wf_title.fill = hfill(C_HDR_DARK)
    wf_title.font = hfont("FFFFFF", bold=True, size=12)
    wf_title.alignment = center()
    wf_title.border = border()
    ws_wf.row_dimensions[1].height = 28

    cur_row = 2
    METRIC_LABELS = {
        "exception-triage": [
            ("Plan ID",                "plan_id"),
            ("Exception Types Found",  ("exception_types", len)),
            ("Decisions Made",         "decisions_made"),
        ],
        "late-supplier": [
            ("Candidates Reviewed",              "candidates_reviewed"),
            ("Chronic Late Suppliers Flagged",
             "chronic_late_suppliers_flagged"),
        ],
        "safety-stock": [
            ("Items Below Safety Stock",         "items_below_safety_stock"),
            ("Items With Excess Inventory",       "items_with_excess_inventory"),
            ("Total Items Reviewed",              "total_items_reviewed"),
        ],
        "price-anomaly": [
            ("Price Anomalies Found",             "anomalies_found"),
            ("Items Over Contract Price",         "items_over_contract_price"),
            ("Items Without Active Contract",     "items_without_active_contract"),
        ],
        "demand-to-po": [
            ("Pegging Records Reviewed",          "pegging_records_reviewed"),
        ],
        "spend-analytics": [
            ("Suppliers Ranked",                  "suppliers_ranked"),
            ("Total Spend (USD)",                 "total_spend"),
            ("Top-3 Concentration %",             "top3_concentration_pct"),
            ("Maverick Spend Items",              "maverick_spend_count"),
            ("Single Source Items",               "single_source_items"),
            ("Consolidation Opportunities",       "consolidation_opportunities"),
            ("Quarterly Periods Analysed",        "quarterly_periods_analysed"),
        ],
    }

    for wf_name, wf_data in workflows.items():
        # Workflow name banner
        ws_wf.merge_cells(f"A{cur_row}:C{cur_row}")
        banner = ws_wf.cell(cur_row, 1, wf_name.upper().replace("-", " "))
        banner.fill = hfill(C_HDR_MID)
        banner.font = hfont("FFFFFF", bold=True)
        banner.alignment = left()
        banner.border = border()
        ws_wf.row_dimensions[cur_row].height = 20
        cur_row += 1

        metric_defs = METRIC_LABELS.get(wf_name, [])
        if not metric_defs:
            # Fallback: list top-level numeric/string keys
            metric_defs = [(k, k) for k in wf_data if not isinstance(
                wf_data[k], (dict, list))]

        for label, key in metric_defs:
            if isinstance(key, tuple):
                field, fn = key
                val = fn(wf_data.get(field, []))
            else:
                val = wf_data.get(key, "—")
            row_fill = C_ROW_ALT if cur_row % 2 == 0 else C_WHITE
            lc = ws_wf.cell(cur_row, 1, label)
            lc.fill = hfill(C_LABEL_BG)
            lc.font = hfont("000000", bold=True, size=10)
            lc.border = border()
            lc.alignment = left()
            vc = ws_wf.cell(cur_row, 2, _safe(val))
            vc.font = Font(name="Arial", size=10)
            vc.fill = hfill(row_fill)
            vc.border = border()
            vc.alignment = left()
            cur_row += 1

        # Actions for this workflow
        wf_acts = [a for a in actions if a.get("workflow") == wf_name]
        if wf_acts:
            hrow = cur_row
            for ci, hdr_label in enumerate(["Action", "Priority", "Summary"], 1):
                hc = ws_wf.cell(hrow, ci, hdr_label)
                hc.fill = hfill(C_HDR_DARK)
                hc.font = hfont("FFFFFF", bold=True, size=9)
                hc.border = border()
                hc.alignment = center()
            cur_row += 1
            for action in wf_acts:
                priority = action.get("priority", "P4-LOW")
                p_fill, p_txt = _priority_colors(priority)
                row_fill = C_ROW_ALT if cur_row % 2 == 0 else C_WHITE
                _data_cell(ws_wf, cur_row, 1, action.get(
                    "action", ""), fill=row_fill)
                _data_cell(ws_wf, cur_row, 2, priority, fill=p_fill,
                           txt_color=p_txt, bold=True, align="center")
                _data_cell(ws_wf, cur_row, 3, action.get(
                    "summary", ""), fill=row_fill, wrap=True)
                ws_wf.row_dimensions[cur_row].height = 36
                cur_row += 1

        cur_row += 1  # blank spacer between workflows

    _set_col_widths(ws_wf, [30, 20, 60])

    wb.save(output_path)
    return output_path


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Standalone tutorial-based Agentic Procurement AI runner.",
    )
    parser.add_argument(
        "--engine",
        choices=["rules", "claude"],
        default="rules",
        help="Execution engine. Use claude for full tutorial-style tool use.",
    )
    parser.add_argument(
        "--workflow",
        choices=[
            "all",
            "exception-triage",
            "late-supplier",
            "safety-stock",
            "price-anomaly",
            "demand-to-po",
            "spend-analytics",
        ],
        default="exception-triage",
        help="Which tutorial workflow to run.",
    )
    parser.add_argument(
        "--autonomy-level",
        type=int,
        choices=[0, 1, 2],
        default=1,
        help=(
            "Tutorial autonomy level. Level 2 emits draft PO payloads; "
            "lower levels recommend only."
        ),
    )
    parser.add_argument(
        "--plan-id",
        type=int,
        default=None,
        help=(
            "Specific MSC plan ID. Defaults to the plan with the most "
            "exceptions."
        ),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=10,
        help="Max records to inspect per workflow.",
    )
    parser.add_argument(
        "--organization-id",
        type=int,
        default=None,
        help=(
            "Optional Oracle inventory organization / operating unit override. "
            "Use this to run the agent for a specific org such as 7088."
        ),
    )
    return parser


def _print_banner() -> None:
    print("=" * 72)
    print("  TUTORIAL AGENTIC PROCUREMENT AI")
    print(
        "  Read-only Oracle analysis | auditable actions | "
        "draft PO simulation"
    )
    print("=" * 72)


def _save_report(report: dict[str, Any]) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = report.get(
        "run_id") or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"tutorial_agent_run_{timestamp}.json"
    output_path.write_text(json.dumps(report, indent=2, default=_json_default))
    return output_path


def run_agent(
    workflow: str = "exception-triage",
    autonomy_level: int = 1,
    plan_id: int | None = None,
    limit: int = 10,
    organization_id: int | None = None,
    engine: str = "rules",
) -> dict[str, Any]:
    gateway = OracleReadOnlyGateway()
    connection_info = gateway.test_connection()

    agent = TutorialProcurementAgent(
        gateway,
        autonomy_level=autonomy_level,
        organization_id=organization_id,
    )
    agent.reset_actions()

    if engine == "claude":
        result = _run_claude_agent(
            gateway=gateway,
            agent=agent,
            workflow=workflow,
            plan_id=plan_id,
            limit=limit,
        )
    else:
        result = agent.run(
            workflow=workflow,
            plan_id=plan_id,
            limit=limit,
        )
        result["engine"] = "rules"
        result["model"] = "deterministic-python-rules"
        result["transcript"] = []
        result["final_text"] = "Rules-based tutorial run complete."

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    report = {
        "run_id": run_id,
        "generated_at": datetime.now().isoformat(),
        "tutorial_source": str(TUTORIAL_PATH.name),
        "connection": connection_info,
        "result": result,
    }
    output_path = _save_report(report)
    excel_path = _save_excel(report)
    return {
        "report": report,
        "json_path": str(output_path),
        "excel_path": str(excel_path),
        "connection": connection_info,
        "result": result,
    }


def main() -> int:
    parser = _build_parser()
    args = parser.parse_args()

    _print_banner()
    print(f"Tutorial source : {TUTORIAL_PATH.name}")
    print(f"Engine          : {args.engine}")
    print(f"Workflow        : {args.workflow}")
    print(f"Autonomy level  : {args.autonomy_level}")
    print(f"Record limit    : {args.limit}")
    print(f"Organization ID : {args.organization_id or 'auto'}")

    try:
        run_output = run_agent(
            workflow=args.workflow,
            autonomy_level=args.autonomy_level,
            plan_id=args.plan_id,
            limit=args.limit,
            organization_id=args.organization_id,
            engine=args.engine,
        )
        report = run_output["report"]
        connection_info = run_output["connection"]
        result = run_output["result"]
        output_path = Path(run_output["json_path"])
        excel_path = Path(run_output["excel_path"])
        print(
            f"Connected to    : {connection_info['sid']} as "
            f"{connection_info['connected_user']} "
            f"(DB version {connection_info['db_version']})"
        )

        print("\nAction summary")
        print("-" * 72)
        for action_name, count in sorted(result["action_summary"].items()):
            print(f"{action_name:20} {count}")
        if not result["action_summary"]:
            print("No actions were produced.")

        print(f"\nSaved report    : {output_path}")
        print(f"Saved workbook  : {excel_path}")
        print("\nSample actions")
        print("-" * 72)
        for action in result["actions"][:10]:
            print(
                f"[{action['workflow']}] {action['action']} :: "
                f"{action['summary']}"
            )
        if result.get("final_text"):
            print("\nAgent summary")
            print("-" * 72)
            print(result["final_text"])

        return 0
    except Exception as exc:
        print("\nAgent run failed")
        print("-" * 72)
        print(str(exc))
        print(traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
